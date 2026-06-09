"""
Intercompany API.

Tracks GL accounts that represent transactions with related entities
(parents, subs, sister companies). For consolidated reporting these
balances must be eliminated against the matching account on the
counterparty's books.

Endpoints:
  GET    /intercompany/overview?period_end=YYYY-MM-DD
         List flagged IC accounts with current-period balance + change
  POST   /intercompany/auto-detect
         Scan every QBO balance-sheet account, mark anything matching
         the IC name pattern. Non-destructive: existing rows stay as-is.
  POST   /intercompany/marks
         Add (or update) a manual IC mark.
  DELETE /intercompany/marks/{id}
         Remove a mark.
  GET    /intercompany/account/{qbo_id}/transactions?period_end=YYYY-MM-DD
         GL transactions for an IC account in the closing month.
"""
import logging
import re
import uuid
from datetime import date
from decimal import Decimal

from fastapi import APIRouter, Depends, HTTPException, Query
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from core.ai.guard import enforce_ai_limits
from core.auth.dependencies import CurrentTenantId, CurrentUser, require_role
from core.db.session import get_db
from models.gl_balance_snapshot import GlBalanceSnapshot
from models.intercompany_account import IntercompanyAccount
from models.intercompany_pair import IntercompanyPair
from models.qbo_connection import QboConnection
from models.tenant import Tenant
from models.user import User

logger = logging.getLogger(__name__)

router = APIRouter()


# ── Schemas ─────────────────────────────────────────────────────────────────

class IcAccountOut(BaseModel):
    id:              str
    qbo_account_id:  str
    account_number:  str
    account_name:    str
    account_type:    str
    counterparty:    str | None
    kind:            str                # 'receivable' | 'payable' | 'unknown'
    auto_detected:   bool
    notes:           str | None
    # Live data from QBO at the requested period_end
    gl_balance:      str
    prior_balance:   str | None         # one month prior, for change calc
    change:          str | None
    created_at:      str
    updated_at:      str


class OverviewResponse(BaseModel):
    qbo_connected:    bool
    period_end:       str
    accounts:         list[IcAccountOut]
    totals:           dict               # receivables, payables, net
    detected_pending: int                # IC-pattern accounts not yet marked


class MarkIn(BaseModel):
    qbo_account_id: str
    counterparty:   str | None = None
    kind:           str = "unknown"      # 'receivable' | 'payable' | 'unknown'
    notes:          str | None = None


# ── Name-pattern auto-detection ──────────────────────────────────────────────

# Case-insensitive substrings — if an account's name contains any of these
# AND its account type is a balance-sheet account, it's likely IC. Tuned to
# catch the common real-world naming conventions (spelled out, abbreviated,
# plural, with/without separators) while avoiding obvious false positives.
#
# Coverage:
#   * Spelled out:   intercompany / inter-company / inter company
#   * Abbreviated:   interco / inter-co / inter co / intco
#   * Due to/from:   "Due to Parent Co", "Due from Sub LLC"
#   * I/C:           i/c, i.c., i.c
#   * IC + noun:     "IC Payable", "IC Receivables" (plural too)
#   * Related party: "Loan from Parent", "Note Receivable - HoldCo",
#                    "Subsidiary Loan", "Affiliate Receivables",
#                    "Investment in Subsidiary", "Holding Co Balance"
#   * Generic:       "Related party"
_IC_PATTERNS = [
    # Spelled out with optional separator (space or hyphen) between "inter"
    # and "company". Matches "Intercompany Payables", "Inter-company Loan",
    # "Inter Company Receivable".
    r"\binter[\s\-]?company\b",
    # Abbreviated form — "interco", "inter-co", "inter co", "intco". Both
    # parts must be word-bounded so we don't fire on "internet co" or
    # "interior".
    r"\binter[\s\-]?co\b",
    r"\bintco\b",
    # Universal IC phrasing — "Due to X", "Due from X". The X is usually
    # the counterparty name (parsed in pass 2).
    r"\bdue\s+(to|from)\b",
    # Slash / dot abbreviations  — "i/c", "i.c.", "i.c"
    r"\bi/c\b",
    r"\bi\.c\.?\b",
    # "IC [payable|receivable|loan|balance|account]" with optional plural
    # — fixes the bug where "IC Payables" wasn't matching "payable\b".
    r"\bic\s+(receivable|payable|loan|balance|account)s?\b",
    # Related-party / affiliate phrasings — most flexible bucket. Catches
    # noun-then-related-entity and related-entity-then-noun, with up to
    # 20 chars of separator (preposition, dash, colon, etc.) between them.
    # Examples it catches:
    #   "Loan from Parent", "Note Receivable - HoldCo",
    #   "Payable to Sister Co", "Investment in Subsidiary"
    #   "Subsidiary Loan", "Affiliate Receivables",
    #   "Parent Payable", "HoldCo Balance"
    r"\b(note|loan|advance|payable|receivable|balance|investment)\b.{0,20}\b(subsidiary|parent|affiliate|sister|holdco|holding|related\s+party)\b",
    r"\b(subsidiary|parent|affiliate|sister|holdco|holding)\b.{0,20}\b(receivable|payable|loan|advance|balance|investment)s?\b",
    # Generic "Related party" catch-all
    r"\brelated\s+part(y|ies)\b",
]
_IC_REGEX = re.compile("|".join(_IC_PATTERNS), re.IGNORECASE)


def _kind_for_account_type(acct_type: str) -> str:
    """Default kind guess from QBO AccountType."""
    if acct_type in ("Accounts Receivable", "Other Current Asset", "Other Asset", "Fixed Asset"):
        return "receivable"
    if acct_type in ("Accounts Payable", "Other Current Liability", "Long Term Liability"):
        return "payable"
    return "unknown"


# ── Endpoints ────────────────────────────────────────────────────────────────

@router.get("/overview", response_model=OverviewResponse)
async def get_overview(
    tenant_id: CurrentTenantId,
    period_end: str = Query(..., description="Period end YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
) -> OverviewResponse:
    """All flagged IC accounts with their live GL balances at period_end."""
    try:
        pe = date.fromisoformat(period_end)
    except ValueError:
        raise HTTPException(400, "period_end must be YYYY-MM-DD.")

    conn = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id == tenant_id),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    if conn is None:
        return OverviewResponse(
            qbo_connected=False,
            period_end=pe.isoformat(),
            accounts=[],
            totals={"receivables": "0.00", "payables": "0.00", "net": "0.00"},
            detected_pending=0,
        )

    # 1. Load every marked IC account for this tenant
    marks = list((await db.execute(select(IntercompanyAccount))).scalars().all())
    mark_by_qbo: dict[str, IntercompanyAccount] = {m.qbo_account_id: m for m in marks}

    # 2. Pull QBO accounts + current and prior trial balances (for change)
    from core.qbo_tb import fetch_trial_balance, lookup_balance, parse_trial_balance
    from modules.recons.overview import _list_balance_sheet_accounts

    accts_raw = await _list_balance_sheet_accounts(conn, db)
    accts_by_id = {str(a.get("Id")): a for a in accts_raw}

    # Trial balances
    try:
        tb_cur_raw = await fetch_trial_balance(conn, pe)
        tb_cur = parse_trial_balance(tb_cur_raw)
    except Exception:
        logger.exception("TB pull failed for IC overview at %s", pe)
        tb_cur = {"by_id": {}, "by_name": {}, "rows": 0}

    # Prior period = last day of prior month
    prior_pe = (pe.replace(day=1) - __import__("datetime").timedelta(days=1))
    try:
        tb_prior_raw = await fetch_trial_balance(conn, prior_pe)
        tb_prior = parse_trial_balance(tb_prior_raw)
    except Exception:
        tb_prior = {"by_id": {}, "by_name": {}, "rows": 0}

    out: list[IcAccountOut] = []
    receivables = Decimal("0")
    payables    = Decimal("0")
    for m in marks:
        a = accts_by_id.get(m.qbo_account_id, {})
        cur_bal = lookup_balance(tb_cur, qbo_id=m.qbo_account_id,
                                  acct_num=str(a.get("AcctNum") or ""), name=str(a.get("Name") or ""))
        prior_bal = lookup_balance(tb_prior, qbo_id=m.qbo_account_id,
                                    acct_num=str(a.get("AcctNum") or ""), name=str(a.get("Name") or ""))
        cur_d   = cur_bal or Decimal("0")
        prior_d = prior_bal if prior_bal is not None else None
        change  = (cur_d - prior_d).quantize(Decimal("0.01")) if prior_d is not None else None

        # Roll-up into totals using kind. Use absolute value: a debit
        # balance on a receivable account is positive, but a "negative"
        # receivable (overdrawn balance) shouldn't reduce the total —
        # better to surface it as its own row anomaly.
        if m.kind == "receivable":
            receivables += abs(cur_d)
        elif m.kind == "payable":
            payables += abs(cur_d)

        out.append(IcAccountOut(
            id=str(m.id),
            qbo_account_id=m.qbo_account_id,
            account_number=str(a.get("AcctNum") or ""),
            account_name=str(a.get("Name") or "(missing in QBO)"),
            account_type=str(a.get("AccountType") or ""),
            counterparty=m.counterparty,
            kind=m.kind,
            auto_detected=m.auto_detected,
            notes=m.notes,
            gl_balance=str(cur_d.quantize(Decimal("0.01"))),
            prior_balance=str(prior_d.quantize(Decimal("0.01"))) if prior_d is not None else None,
            change=str(change) if change is not None else None,
            created_at=m.created_at.isoformat(),
            updated_at=m.updated_at.isoformat(),
        ))

    # 3. Count IC-pattern accounts that AREN'T yet marked — for the
    # "Auto-detect found N candidates" CTA on an empty/partial dashboard.
    detected_pending = 0
    for acct in accts_raw:
        qid = str(acct.get("Id") or "")
        if not qid or qid in mark_by_qbo:
            continue
        if _IC_REGEX.search(str(acct.get("Name") or "")):
            detected_pending += 1

    net = (receivables - payables).quantize(Decimal("0.01"))
    return OverviewResponse(
        qbo_connected=True,
        period_end=pe.isoformat(),
        accounts=out,
        totals={
            "receivables": str(receivables.quantize(Decimal("0.01"))),
            "payables":    str(payables.quantize(Decimal("0.01"))),
            "net":         str(net),
        },
        detected_pending=detected_pending,
    )


@router.post("/auto-detect")
async def auto_detect(
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
    classify: bool = True,
) -> dict:
    """
    Two-pass scan that does most of the IC setup work automatically:
      1. Name-pattern detection — flag accounts whose names look like
         IC (Due to/from, Intercompany, etc.).
      2. (when classify=true) Auto-fill counterparty from the modal
         entity name across the last 6 months of transactions per
         newly-detected account. Skips accounts whose transactions
         span multiple entities (under 50% modal).
      3. Counterparty inference also tries to read it from the
         account NAME itself when the txn pull yields nothing —
         e.g. "Due from Acme Sub LLC" → counterparty="Acme Sub LLC".

    Run on every QBO sync (cheap — names rarely change) or
    on-demand via the Intercompany page.
    """
    from collections import Counter
    from datetime import date as _date
    from datetime import timedelta as _td

    from core.qbo_gl import pull_gl_transactions
    from modules.recons.overview import _list_balance_sheet_accounts

    conn = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id == tenant_id),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    if conn is None:
        raise HTTPException(409, "QuickBooks isn't connected.")

    accts = await _list_balance_sheet_accounts(conn, db)
    existing = list((await db.execute(select(IntercompanyAccount))).scalars().all())
    existing_ids = {m.qbo_account_id for m in existing}

    today = _date.today()
    period_start = (today - _td(days=6 * 31)).replace(day=1)

    added = 0
    classified = 0
    # Diagnostic counters — surfaced in the response so the user can see
    # what the scanner did. Critical for debugging "why didn't it find my
    # account?" — if scanned=0 the QBO query failed; if scanned>0 and
    # matched=0 the naming doesn't fit any pattern.
    scanned = len(accts)
    already_marked = 0
    matched_total = 0
    skipped_examples: list[str] = []
    new_marks: list[IntercompanyAccount] = []
    for a in accts:
        qid = str(a.get("Id") or "")
        name = str(a.get("Name") or "")
        if not qid:
            continue
        if qid in existing_ids:
            already_marked += 1
            continue
        if not _IC_REGEX.search(name):
            # Keep up to 10 non-matching names for diagnostic purposes
            if len(skipped_examples) < 10:
                skipped_examples.append(name)
            continue
        matched_total += 1

        # Heuristic counterparty from the account name: strip the
        # "Due to / Due from / Intercompany — " prefix and take the rest.
        cp_from_name: str | None = None
        cleaned = re.sub(
            r"^(intercompany|inter-?company|due\s+to|due\s+from|i\.?c\.?|loan\s+(?:to|from))[\s\-:]*",
            "", name, flags=re.IGNORECASE,
        ).strip(" -—:")
        if cleaned and len(cleaned) >= 2 and cleaned.lower() != name.lower():
            cp_from_name = cleaned

        row = IntercompanyAccount(
            qbo_account_id=qid,
            counterparty=cp_from_name,
            kind=_kind_for_account_type(str(a.get("AccountType") or "")),
            auto_detected=True,
            created_by=user.id,
        )
        db.add(row)
        new_marks.append(row)
        added += 1
        if cp_from_name:
            classified += 1

    # Flush so new rows have IDs before we query against them
    await db.flush()

    # Pass 2 — transaction-derived counterparty for any new marks that
    # didn't get one from the name heuristic.
    if classify:
        for m in new_marks:
            if m.counterparty:
                continue
            try:
                rows = await pull_gl_transactions(conn, db, m.qbo_account_id, period_start, today)
            except Exception:
                continue
            names = [(r["entity_name"] or "").strip() for r in rows]
            names = [n for n in names if n]
            if not names:
                continue
            top = Counter(names).most_common(1)[0]
            if top[1] / len(rows) >= 0.5:
                m.counterparty = top[0]
                classified += 1

    await db.commit()
    return {
        "added":          added,
        "classified":     classified,
        "scanned":        scanned,
        "matched":        matched_total,
        "already_marked": already_marked,
        # First few unmatched names — gives the user a hint about whether
        # the scanner ran successfully but their naming convention isn't
        # in the pattern list yet.
        "skipped_sample": skipped_examples[:5],
    }


@router.post("/marks")
async def upsert_mark(
    body: MarkIn,
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    if body.kind not in ("receivable", "payable", "unknown"):
        raise HTTPException(400, "kind must be receivable / payable / unknown")

    existing = (await db.execute(
        select(IntercompanyAccount).where(IntercompanyAccount.qbo_account_id == body.qbo_account_id)
    )).scalar_one_or_none()
    if existing is None:
        existing = IntercompanyAccount(
            qbo_account_id=body.qbo_account_id,
            counterparty=body.counterparty,
            kind=body.kind,
            auto_detected=False,
            notes=body.notes,
            created_by=user.id,
        )
        db.add(existing)
    else:
        existing.counterparty = body.counterparty
        existing.kind = body.kind
        existing.notes = body.notes
        # Once a user touches the row, it's no longer purely auto-detected.
        existing.auto_detected = False
    await db.commit()
    await db.refresh(existing)
    return {"id": str(existing.id), "ok": True}


@router.delete("/marks/{mark_id}", dependencies=[Depends(require_role("reviewer"))])
async def delete_mark(
    mark_id: uuid.UUID,
    tenant_id: CurrentTenantId,
    db: AsyncSession = Depends(get_db),
) -> dict:
    row = (await db.execute(
        select(IntercompanyAccount).where(IntercompanyAccount.id == mark_id)
    )).scalar_one_or_none()
    if row is None:
        raise HTTPException(404, "Mark not found.")
    await db.delete(row)
    await db.commit()
    return {"ok": True}


@router.post("/ai-detect", dependencies=[Depends(enforce_ai_limits)])
async def ai_detect(
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """
    AI-powered IC detection — pass the whole chart of accounts to Claude
    and let it identify intercompany accounts that the regex auto-detect
    missed. Useful for non-standard naming ("Loan – HoldCo", "Owner
    Investment – Sub", "Mgmt Fee Receivable - Parent", etc).

    Marks any account with AI confidence ≥ 0.6, auto_detected=True, and
    writes the AI's reasoning into the notes field. Existing marks
    aren't touched. Returns summary so the UI can render results.
    """
    import json as _json

    import anthropic as _anthropic

    from core.config import settings as _settings
    from modules.recons.overview import _list_balance_sheet_accounts

    conn = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id == tenant_id),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    if conn is None:
        raise HTTPException(409, "QuickBooks isn't connected.")

    accts = await _list_balance_sheet_accounts(conn, db)
    existing = list((await db.execute(select(IntercompanyAccount))).scalars().all())
    existing_ids = {m.qbo_account_id for m in existing}

    # Only send accounts that aren't already marked — saves tokens and
    # avoids reclassifying. If the user wants to re-evaluate an existing
    # mark they can delete it first.
    scan_pool = [
        {
            "id":   str(a.get("Id") or ""),
            "name": str(a.get("Name") or ""),
            "type": str(a.get("AccountType") or ""),
            "num":  str(a.get("AcctNum") or ""),
        }
        for a in accts
        if str(a.get("Id") or "") and str(a.get("Id") or "") not in existing_ids
    ]

    if not scan_pool:
        return {
            "added":           0,
            "scanned":         len(accts),
            "ai_candidates":   0,
            "already_marked":  len(existing),
            "skipped_lowconf": 0,
        }

    # ── Build the prompt ──────────────────────────────────────────────
    system_prompt = """\
You are a senior accountant reviewing a company's Chart of Accounts.

Your job: identify GL accounts that record INTERCOMPANY (related party)
transactions — balances with a parent, subsidiary, sister entity,
affiliate, or owner-controlled related party. These accounts are
typically eliminated on consolidation.

Common signals:
  * Names containing: Intercompany, Inter-co, I/C, IC, Due to/from,
    Affiliate, Subsidiary, Parent, HoldCo, Holding, Sister Co,
    Related party
  * "Investment in [SubName]", "Note Receivable – [EntityName]",
    "Loan to/from [EntityName]" where the entity is a related party
  * Owner / member loan / advance accounts in closely held entities
  * "Management Fee Receivable/Payable" where it's an inter-entity flow
  * Account names containing what looks like another LLC / Inc name
    when paired with words like Loan, Payable, Receivable, Investment

NOT intercompany:
  * Customer A/R, Trade A/P (unless name explicitly says related party)
  * Bank, Credit Card, Fixed Assets (unless name signals IC)
  * Operating Expense / Income / standard equity accounts
  * Generic accruals, prepaids without an entity name

You will respond with a single JSON object:
{
  "matches": [
    {
      "id":            "<account id from input>",
      "confidence":    <0.0 to 1.0 — how sure>,
      "kind":          "<receivable | payable | unknown>",
      "counterparty":  "<entity name extracted from account name, or null>",
      "reason":        "<one short sentence>"
    }
  ]
}

Be conservative. Confidence ≥ 0.6 means you would flag this account in
a real audit prep. Confidence < 0.6 means it's a maybe — leave it out.
Empty list is acceptable.
"""

    lines = []
    for i, a in enumerate(scan_pool):
        lines.append(
            f'{i}. id="{a["id"]}" num="{a["num"]}" name="{a["name"]}" type="{a["type"]}"'
        )
    user_msg = (
        f"Review these {len(scan_pool)} accounts and return your "
        f"JSON matches array.\n\n" + "\n".join(lines)
    )

    client = _anthropic.Anthropic(api_key=_settings.anthropic_api_key)
    try:
        resp = client.messages.create(
            model=_settings.anthropic_model,
            max_tokens=4000,
            system=system_prompt,
            messages=[{"role": "user", "content": user_msg}],
        )
    except Exception:
        logger.exception("IC ai-detect: Claude call failed")
        raise HTTPException(502, "AI scan failed. Try again in a moment.")

    from core.ai.usage import record_response
    record_response(resp, operation="intercompany_detect")

    text = "".join(
        getattr(b, "text", "") for b in resp.content
        if getattr(b, "type", None) == "text"
    )
    text = re.sub(r"^```(?:json)?\s*|\s*```\s*$", "", text.strip(),
                  flags=re.IGNORECASE | re.MULTILINE)

    try:
        parsed = _json.loads(text)
        matches = parsed.get("matches") if isinstance(parsed, dict) else None
        if not isinstance(matches, list):
            matches = []
    except Exception:
        logger.exception("IC ai-detect: JSON parse failed (text head): %s", text[:300])
        matches = []

    # Build a quick lookup from id back to its source account so we can
    # set kind defaults and look up the name for notes.
    pool_by_id = {a["id"]: a for a in scan_pool}

    added = 0
    skipped_lowconf = 0
    for m in matches:
        if not isinstance(m, dict):
            continue
        qid = str(m.get("id") or "")
        if not qid or qid not in pool_by_id or qid in existing_ids:
            continue
        try:
            conf = float(m.get("confidence") or 0)
        except Exception:
            conf = 0.0
        if conf < 0.6:
            skipped_lowconf += 1
            continue

        kind = str(m.get("kind") or "").lower()
        if kind not in ("receivable", "payable", "unknown"):
            kind = _kind_for_account_type(pool_by_id[qid]["type"])

        cp = str(m.get("counterparty") or "").strip() or None
        reason = str(m.get("reason") or "").strip()
        notes = f"AI-detected (conf {conf:.0%}): {reason}" if reason else None

        db.add(IntercompanyAccount(
            qbo_account_id=qid,
            counterparty=cp,
            kind=kind,
            auto_detected=True,
            notes=notes,
            created_by=user.id,
        ))
        existing_ids.add(qid)
        added += 1

    await db.commit()
    return {
        "added":           added,
        "scanned":         len(accts),
        "ai_candidates":   len(matches),
        "already_marked":  len(existing),
        "skipped_lowconf": skipped_lowconf,
    }


@router.post("/auto-classify")
async def auto_classify(
    tenant_id: CurrentTenantId,
    db: AsyncSession = Depends(get_db),
    lookback_months: int = 6,
) -> dict:
    """
    For every marked IC account that doesn't already have a
    counterparty filled in, scan recent transactions and use the
    most-common `entity_name` as the counterparty. This dramatically
    reduces the amount of manual classification work — most IC
    accounts only ever transact with ONE related entity, so
    transaction-name → counterparty is reliable.

    Idempotent: only fills empty counterparty fields. Doesn't overwrite
    user-entered values.
    """
    from collections import Counter
    from datetime import date as _date
    from datetime import timedelta as _td

    from core.qbo_gl import pull_gl_transactions

    conn = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id == tenant_id),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    if conn is None:
        raise HTTPException(409, "QuickBooks isn't connected.")

    today = _date.today()
    period_start = (today - _td(days=lookback_months * 31)).replace(day=1)

    marks = list((await db.execute(
        select(IntercompanyAccount).where(IntercompanyAccount.counterparty.is_(None))
    )).scalars().all())

    classified = 0
    for m in marks:
        try:
            rows = await pull_gl_transactions(conn, db, m.qbo_account_id, period_start, today)
        except Exception:
            logger.exception("auto-classify: GL pull failed for %s", m.qbo_account_id)
            continue
        if not rows:
            continue
        # Tally entity_name occurrences; pick the modal value if it
        # accounts for >= 50% of transactions (otherwise the account
        # transacts with multiple parties and we leave it blank for
        # the user to decide).
        names = [(r["entity_name"] or "").strip() for r in rows]
        names = [n for n in names if n]
        if not names:
            continue
        c = Counter(names).most_common(1)[0]
        winner, votes = c[0], c[1]
        if votes / len(rows) >= 0.5:
            m.counterparty = winner
            classified += 1
    await db.commit()
    return {"classified": classified, "considered": len(marks)}


@router.get("/account/{qbo_account_id}/transactions")
async def get_account_transactions(
    qbo_account_id: str,
    tenant_id: CurrentTenantId,
    period_end: str = Query(..., description="Period end YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Transactions posted to this IC account in the closing month."""
    try:
        pe = date.fromisoformat(period_end)
    except ValueError:
        raise HTTPException(400, "period_end must be YYYY-MM-DD.")
    conn = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id == tenant_id),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    if conn is None:
        raise HTTPException(409, "QuickBooks isn't connected.")

    from core.qbo_gl import pull_gl_transactions
    period_start = pe.replace(day=1)
    rows = await pull_gl_transactions(conn, db, qbo_account_id, period_start, pe)

    out = []
    total = Decimal("0")
    for r in rows:
        total += r["amount"]
        out.append({
            "txn_id":   r["qbo_txn_id"] or "",
            "txn_type": r["txn_type"],
            "txn_number": r["txn_number"] or "",
            "txn_date": r["txn_date"].isoformat() if r["txn_date"] else "",
            "amount":   str(r["amount"]),
            "memo":     r["memo"] or "",
            "entity":   r["entity_name"] or "",
        })
    return {
        "rows": out,
        "total": str(total.quantize(Decimal("0.01"))),
        "period_start": period_start.isoformat(),
        "period_end":   pe.isoformat(),
    }


# ═══════════════════════════════════════════════════════════════════════════
# Cross-tenant pairing + eliminations + consolidated TB
# ═══════════════════════════════════════════════════════════════════════════
#
# These endpoints turn the IC module from "tracking" into actual
# consolidation prep. A user who is a member of multiple Clerk orgs
# (typical for a CPA managing a multi-entity client group) can:
#
#   1. PAIR an IC account in one workspace with the matching account in
#      another workspace they have access to.  /api/intercompany/pairs
#
#   2. See an ELIMINATIONS report — for every pair, both sides' balances
#      at period_end + the diff. Matched pairs eliminate; mismatched
#      pairs flag for investigation.   /api/intercompany/eliminations
#
#   3. Get a CONSOLIDATED TB across every linked entity with the
#      eliminations applied. Exportable to Excel for working-paper use.
#                                       /api/intercompany/consolidated-tb
#
# Cross-tenant safety: every query that crosses tenant boundaries first
# verifies the current user is a member of the destination tenant via
# the User table (same clerk_user_id → row exists in other tenant). The
# row-level filter is bypassed via skip_tenant_filter ONLY when the
# membership check has succeeded.


# ── Schemas ─────────────────────────────────────────────────────────────────

class AccessibleAccount(BaseModel):
    qbo_account_id: str
    account_number: str
    account_name:   str
    account_type:   str
    kind:           str   # 'receivable' | 'payable' | 'unknown'
    counterparty:   str | None


class AccessibleCompany(BaseModel):
    tenant_id:       str
    clerk_org_id:    str
    name:            str           # display name from Tenant.name
    company_name:    str | None    # from QboConnection.company_name
    qbo_connected:   bool
    ic_accounts:     list[AccessibleAccount]


class AccessibleCompaniesResponse(BaseModel):
    companies: list[AccessibleCompany]


class PairIn(BaseModel):
    my_qbo_account_id:           str
    counterparty_tenant_id:      str    # UUID string
    counterparty_qbo_account_id: str
    notes:                       str | None = None


class PairOut(BaseModel):
    pair_group_id:               str
    my_qbo_account_id:           str
    my_account_label:            str
    counterparty_tenant_id:      str
    counterparty_clerk_org_id:   str
    counterparty_label:          str
    counterparty_qbo_account_id: str
    notes:                       str | None
    created_at:                  str


class EliminationRow(BaseModel):
    pair_group_id:               str
    my_qbo_account_id:           str
    my_account_label:            str
    my_balance:                  str
    counterparty_tenant_id:      str
    counterparty_label:          str
    counterparty_balance:        str
    # |my_balance + counterparty_balance| in absolute terms — for an IC
    # AR ↔ IC AP pair the books should net to zero (one debit, one
    # credit). diff = my + cp; if ~0 → matched, else mismatch.
    diff:                        str
    status:                      str   # 'matched' | 'mismatch' | 'one_side_missing'


class EliminationsResponse(BaseModel):
    period_end:    str
    rows:          list[EliminationRow]
    totals: dict   # matched_count, mismatch_count, total_to_eliminate


class ConsolidatedRow(BaseModel):
    fs_category:        str
    account_label:      str          # combined label for the line
    tenant_id:          str
    company_name:       str
    qbo_account_id:     str
    raw_balance:        str
    elimination:        str          # signed amount we subtract
    consolidated:       str          # raw + elimination
    is_eliminated_row:  bool         # true if any part of the balance was eliminated


class ConsolidatedTbResponse(BaseModel):
    period_end:        str
    companies:         list[dict]   # [{ tenant_id, name }]
    rows:              list[ConsolidatedRow]
    totals: dict   # raw_total, elim_total, consolidated_total per category
    # Integrity (Phase 3 trust sweep): a balanced consolidation nets to ~0
    # (debit-positive). balanced=False → out of balance by `imbalance`.
    balanced:          bool = True
    imbalance:         str = "0.00"
    # IC balances that couldn't be eliminated (one side missing or a mismatch),
    # so they still inflate the consolidation. Each: {account_label,
    # company_name, my_balance, counterparty_balance, reason}.
    unmatched:         list[dict] = []


# ── Helpers — cross-tenant access ───────────────────────────────────────────

async def _user_accessible_tenant_ids(db: AsyncSession, user: User) -> set[uuid.UUID]:
    """
    Every tenant the current user has a User row in (i.e. is a Clerk-org
    member of). Bypasses the row-level tenant filter intentionally —
    membership IS the cross-tenant authorization.
    """
    rows = (await db.execute(
        select(User.tenant_id).where(User.clerk_user_id == user.clerk_user_id),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all()
    return {r for r in rows}


async def _ensure_user_can_access(db: AsyncSession, user: User, target_tenant_id: uuid.UUID) -> None:
    """403 unless the user has membership in `target_tenant_id`."""
    accessible = await _user_accessible_tenant_ids(db, user)
    if target_tenant_id not in accessible:
        raise HTTPException(403, "You don't have access to that workspace.")


async def _account_label_from_snapshot(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    qbo_account_id: str,
) -> str:
    """
    Build a display label "AccountNum AccountName" for an IC account in
    a given tenant from the latest GL snapshot. Returns "(account)"
    when no snapshot exists yet (recons hasn't been synced).
    """
    row = (await db.execute(
        select(GlBalanceSnapshot)
        .where(
            GlBalanceSnapshot.tenant_id == tenant_id,
            GlBalanceSnapshot.qbo_account_id == qbo_account_id,
        )
        .order_by(GlBalanceSnapshot.period_end.desc())
        .limit(1),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    if row is None:
        return "(account)"
    num = (row.account_number or "").strip()
    name = (row.account_name or "").strip()
    return f"{num} {name}".strip() if num else name


async def _balance_at_period_end(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    qbo_account_id: str,
    period_end: date,
) -> Decimal | None:
    """
    Look up the snapshot for this account at period_end (closest <= match
    so a snapshot taken mid-month still serves). Returns None when no
    snapshot exists for this period.
    """
    row = (await db.execute(
        select(GlBalanceSnapshot)
        .where(
            GlBalanceSnapshot.tenant_id == tenant_id,
            GlBalanceSnapshot.qbo_account_id == qbo_account_id,
            GlBalanceSnapshot.period_end <= period_end,
        )
        .order_by(GlBalanceSnapshot.period_end.desc())
        .limit(1),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    if row is None:
        return None
    return Decimal(row.balance)


# ── /accessible-companies ──────────────────────────────────────────────────

@router.get("/accessible-companies", response_model=AccessibleCompaniesResponse)
async def list_accessible_companies(
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> AccessibleCompaniesResponse:
    """
    Every OTHER workspace the current user has access to (via shared
    Clerk org membership), plus each workspace's existing IC accounts.
    Feeds the "Pair with company X account Y" dropdown.
    """
    accessible = await _user_accessible_tenant_ids(db, user)
    accessible.discard(tenant_id)
    if not accessible:
        return AccessibleCompaniesResponse(companies=[])

    # Pull tenants + their QBO connections + their IC accounts in three
    # bulk queries to keep this fast. All use skip_tenant_filter since
    # we're crossing tenant boundaries (authorization already verified).
    tenants = (await db.execute(
        select(Tenant).where(Tenant.id.in_(accessible)),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all()

    conns = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id.in_(accessible)),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all()
    conn_by_tid = {c.tenant_id: c for c in conns}

    ic_marks = (await db.execute(
        select(IntercompanyAccount).where(IntercompanyAccount.tenant_id.in_(accessible)),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all()
    marks_by_tid: dict[uuid.UUID, list[IntercompanyAccount]] = {}
    for m in ic_marks:
        marks_by_tid.setdefault(m.tenant_id, []).append(m)

    # Hydrate account_name/number from latest snapshot per qbo_account_id.
    # One bulk fetch across all needed (tenant_id, qbo_account_id) pairs.
    snap_rows = (await db.execute(
        select(GlBalanceSnapshot).where(
            GlBalanceSnapshot.tenant_id.in_(accessible),
        ),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all()
    # Build dict keyed by (tenant_id, qbo_account_id) → latest snapshot
    latest_snap: dict[tuple[uuid.UUID, str], GlBalanceSnapshot] = {}
    for s in snap_rows:
        key = (s.tenant_id, s.qbo_account_id)
        prior = latest_snap.get(key)
        if prior is None or s.period_end > prior.period_end:
            latest_snap[key] = s

    out: list[AccessibleCompany] = []
    for t in tenants:
        conn = conn_by_tid.get(t.id)
        accts_out: list[AccessibleAccount] = []
        for m in marks_by_tid.get(t.id, []):
            snap = latest_snap.get((t.id, m.qbo_account_id))
            accts_out.append(AccessibleAccount(
                qbo_account_id=m.qbo_account_id,
                account_number=(snap.account_number if snap else "") or "",
                account_name=(snap.account_name if snap else "(account)"),
                account_type=(snap.account_type if snap else ""),
                kind=m.kind,
                counterparty=m.counterparty,
            ))
        # Sort accounts by number then name
        accts_out.sort(key=lambda a: (a.account_number, a.account_name))
        out.append(AccessibleCompany(
            tenant_id=str(t.id),
            clerk_org_id=t.clerk_org_id,
            name=t.name,
            company_name=(conn.company_name if conn else None),
            qbo_connected=conn is not None,
            ic_accounts=accts_out,
        ))
    # Sort companies alphabetically by display name
    out.sort(key=lambda c: c.name.lower())
    return AccessibleCompaniesResponse(companies=out)


# ── /pairs ──────────────────────────────────────────────────────────────────

@router.get("/pairs", response_model=list[PairOut])
async def list_pairs(
    tenant_id: CurrentTenantId,
    db: AsyncSession = Depends(get_db),
) -> list[PairOut]:
    """Pairs visible to THIS tenant (its half of each pair_group)."""
    pairs = list((await db.execute(
        select(IntercompanyPair).order_by(IntercompanyPair.created_at.desc())
    )).scalars().all())
    out: list[PairOut] = []
    for p in pairs:
        my_label = await _account_label_from_snapshot(db, tenant_id, p.my_qbo_account_id)
        out.append(PairOut(
            pair_group_id=str(p.pair_group_id),
            my_qbo_account_id=p.my_qbo_account_id,
            my_account_label=my_label,
            counterparty_tenant_id=str(p.counterparty_tenant_id),
            counterparty_clerk_org_id=p.counterparty_clerk_org_id,
            counterparty_label=p.counterparty_label,
            counterparty_qbo_account_id=p.counterparty_qbo_account_id,
            notes=p.notes,
            created_at=p.created_at.isoformat(),
        ))
    return out


@router.post("/pairs", response_model=PairOut)
async def create_pair(
    body: PairIn,
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> PairOut:
    """
    Create a pair: writes TWO rows (one per tenant) sharing a
    pair_group_id. Idempotent: returns the existing pair if one already
    exists with the same (my_account, cp_tenant, cp_account).
    """
    try:
        cp_tenant_uuid = uuid.UUID(body.counterparty_tenant_id)
    except Exception:
        raise HTTPException(400, "counterparty_tenant_id must be a UUID.")

    # Authorize: user must have access to the counterparty tenant too
    await _ensure_user_can_access(db, user, cp_tenant_uuid)

    # Idempotency check — same triple already paired?
    existing = (await db.execute(
        select(IntercompanyPair).where(
            IntercompanyPair.my_qbo_account_id == body.my_qbo_account_id,
            IntercompanyPair.counterparty_tenant_id == cp_tenant_uuid,
            IntercompanyPair.counterparty_qbo_account_id == body.counterparty_qbo_account_id,
        )
    )).scalar_one_or_none()
    if existing is not None:
        my_label = await _account_label_from_snapshot(db, tenant_id, existing.my_qbo_account_id)
        return PairOut(
            pair_group_id=str(existing.pair_group_id),
            my_qbo_account_id=existing.my_qbo_account_id,
            my_account_label=my_label,
            counterparty_tenant_id=str(existing.counterparty_tenant_id),
            counterparty_clerk_org_id=existing.counterparty_clerk_org_id,
            counterparty_label=existing.counterparty_label,
            counterparty_qbo_account_id=existing.counterparty_qbo_account_id,
            notes=existing.notes,
            created_at=existing.created_at.isoformat(),
        )

    # Look up the counterparty tenant's display info so we can build the
    # cached label for both halves of the pair.
    cp_tenant = (await db.execute(
        select(Tenant).where(Tenant.id == cp_tenant_uuid),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    if cp_tenant is None:
        raise HTTPException(404, "Counterparty workspace not found.")

    cp_conn = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id == cp_tenant_uuid),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    cp_company_name = (cp_conn.company_name if cp_conn else None) or cp_tenant.name

    # Look up the current tenant info for the mirror row's label
    my_tenant = (await db.execute(
        select(Tenant).where(Tenant.id == tenant_id),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    my_conn = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id == tenant_id),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()
    my_company_name = (my_conn.company_name if my_conn else None) or (my_tenant.name if my_tenant else "Workspace")

    # Account labels from snapshots
    cp_account_label = await _account_label_from_snapshot(db, cp_tenant_uuid, body.counterparty_qbo_account_id)
    my_account_label = await _account_label_from_snapshot(db, tenant_id, body.my_qbo_account_id)

    # Build both halves
    pair_group = uuid.uuid4()

    my_side = IntercompanyPair(
        tenant_id=tenant_id,
        pair_group_id=pair_group,
        my_qbo_account_id=body.my_qbo_account_id,
        counterparty_tenant_id=cp_tenant_uuid,
        counterparty_clerk_org_id=cp_tenant.clerk_org_id,
        counterparty_qbo_account_id=body.counterparty_qbo_account_id,
        counterparty_label=f"{cp_company_name} · {cp_account_label}",
        notes=body.notes,
        created_by=user.id,
    )

    cp_side = IntercompanyPair(
        tenant_id=cp_tenant_uuid,
        pair_group_id=pair_group,
        my_qbo_account_id=body.counterparty_qbo_account_id,
        counterparty_tenant_id=tenant_id,
        counterparty_clerk_org_id=(my_tenant.clerk_org_id if my_tenant else ""),
        counterparty_qbo_account_id=body.my_qbo_account_id,
        counterparty_label=f"{my_company_name} · {my_account_label}",
        notes=body.notes,
        created_by=user.id,
    )

    db.add(my_side)
    db.add(cp_side)
    await db.commit()
    await db.refresh(my_side)

    return PairOut(
        pair_group_id=str(my_side.pair_group_id),
        my_qbo_account_id=my_side.my_qbo_account_id,
        my_account_label=my_account_label,
        counterparty_tenant_id=str(my_side.counterparty_tenant_id),
        counterparty_clerk_org_id=my_side.counterparty_clerk_org_id,
        counterparty_label=my_side.counterparty_label,
        counterparty_qbo_account_id=my_side.counterparty_qbo_account_id,
        notes=my_side.notes,
        created_at=my_side.created_at.isoformat(),
    )


@router.delete("/pairs/{pair_group_id}")
async def delete_pair(
    pair_group_id: uuid.UUID,
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    db: AsyncSession = Depends(get_db),
) -> dict:
    """Delete BOTH halves of the pair_group (across tenants)."""
    # Look up both halves (skip tenant filter — pair belongs to two tenants)
    halves = list((await db.execute(
        select(IntercompanyPair).where(IntercompanyPair.pair_group_id == pair_group_id),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all())
    if not halves:
        raise HTTPException(404, "Pair not found.")

    # Authorize: at least one half must be in a tenant the user has access to
    accessible = await _user_accessible_tenant_ids(db, user)
    if not any(h.tenant_id in accessible for h in halves):
        raise HTTPException(403, "You don't have access to this pair.")

    for h in halves:
        await db.delete(h)
    await db.commit()
    return {"ok": True, "deleted": len(halves)}


# ── /eliminations ──────────────────────────────────────────────────────────

@router.get("/eliminations", response_model=EliminationsResponse)
async def get_eliminations(
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    period_end: str = Query(..., description="Period end YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
) -> EliminationsResponse:
    """
    For every pair this tenant participates in: pull both sides'
    period-end balance and compute the elimination diff.

    Convention: balances are stored signed (debit-positive) in the
    snapshot. An IC AR (asset, debit) on Entity A should equal the
    NEGATIVE of the IC AP (liability, credit) on Entity B. So:
        diff = my_bal + cp_bal
    A "matched" pair has diff ≈ 0. A "mismatch" has |diff| > 1.
    """
    try:
        pe = date.fromisoformat(period_end)
    except ValueError:
        raise HTTPException(400, "period_end must be YYYY-MM-DD.")

    # Pull all pairs visible to this tenant (just our half is enough)
    pairs = list((await db.execute(
        select(IntercompanyPair)
    )).scalars().all())

    rows_out: list[EliminationRow] = []
    matched = 0
    mismatch = 0
    total_to_elim = Decimal("0")

    # Defensive: skip pairs where user lost cross-tenant access
    accessible = await _user_accessible_tenant_ids(db, user)

    for p in pairs:
        if p.counterparty_tenant_id not in accessible:
            continue

        my_bal = await _balance_at_period_end(db, tenant_id, p.my_qbo_account_id, pe)
        cp_bal = await _balance_at_period_end(db, p.counterparty_tenant_id, p.counterparty_qbo_account_id, pe)

        my_label = await _account_label_from_snapshot(db, tenant_id, p.my_qbo_account_id)

        if my_bal is None or cp_bal is None:
            status_str = "one_side_missing"
            diff = my_bal or Decimal("0")  # whichever side we have
            if cp_bal is not None:
                diff = (my_bal or Decimal("0")) + cp_bal
            mismatch += 1
        else:
            diff = (my_bal + cp_bal).quantize(Decimal("0.01"))
            # $1 tolerance to absorb rounding from QBO
            if abs(diff) <= Decimal("1.00"):
                status_str = "matched"
                matched += 1
                total_to_elim += abs(my_bal)
            else:
                status_str = "mismatch"
                mismatch += 1

        rows_out.append(EliminationRow(
            pair_group_id=str(p.pair_group_id),
            my_qbo_account_id=p.my_qbo_account_id,
            my_account_label=my_label,
            my_balance=str((my_bal or Decimal("0")).quantize(Decimal("0.01"))),
            counterparty_tenant_id=str(p.counterparty_tenant_id),
            counterparty_label=p.counterparty_label,
            counterparty_balance=str((cp_bal or Decimal("0")).quantize(Decimal("0.01"))),
            diff=str(diff.quantize(Decimal("0.01"))),
            status=status_str,
        ))

    return EliminationsResponse(
        period_end=pe.isoformat(),
        rows=rows_out,
        totals={
            "matched_count":     matched,
            "mismatch_count":    mismatch,
            "total_to_eliminate": str(total_to_elim.quantize(Decimal("0.01"))),
        },
    )


# ── /consolidated-tb ───────────────────────────────────────────────────────

@router.get("/consolidated-tb", response_model=ConsolidatedTbResponse)
async def get_consolidated_tb(
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    period_end: str = Query(..., description="Period end YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
) -> ConsolidatedTbResponse:
    """
    Roll up the GL snapshot from every paired entity at period_end,
    apply elimination amounts to paired IC accounts, return a single
    consolidated view.

    Scope: only the current tenant + tenants this tenant has at least
    one pair with (so we don't accidentally consolidate unrelated
    workspaces the user happens to also be a member of).
    """
    try:
        pe = date.fromisoformat(period_end)
    except ValueError:
        raise HTTPException(400, "period_end must be YYYY-MM-DD.")

    # Collect the set of tenants in this consolidation = current tenant + its pair counterparties.
    pairs = list((await db.execute(select(IntercompanyPair))).scalars().all())
    accessible = await _user_accessible_tenant_ids(db, user)
    cp_tenants = {p.counterparty_tenant_id for p in pairs if p.counterparty_tenant_id in accessible}
    consol_tenants: set[uuid.UUID] = {tenant_id} | cp_tenants

    # Pull tenant display info
    tenants = (await db.execute(
        select(Tenant).where(Tenant.id.in_(consol_tenants)),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all()
    tenant_by_id = {t.id: t for t in tenants}

    conns = (await db.execute(
        select(QboConnection).where(QboConnection.tenant_id.in_(consol_tenants)),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all()
    conn_by_tid = {c.tenant_id: c for c in conns}

    def company_name(tid: uuid.UUID) -> str:
        """Display label for a tenant — prefer the QBO company name, fall
        back to the workspace name, then a generic placeholder."""
        c = conn_by_tid.get(tid)
        if c is not None and c.company_name:
            return c.company_name
        t = tenant_by_id.get(tid)
        return t.name if t is not None else "Workspace"

    # Build elimination amount map per (tenant_id, qbo_account_id):
    # for each matched pair, the amount we eliminate from each side.
    # Convention: eliminate the absolute balance from BOTH sides (so
    # the IC AR on A nets to 0 and the IC AP on B nets to 0).
    elim_by_acct: dict[tuple[uuid.UUID, str], Decimal] = {}
    unmatched: list[dict] = []
    for p in pairs:
        if p.counterparty_tenant_id not in accessible:
            continue
        my_bal = await _balance_at_period_end(db, tenant_id, p.my_qbo_account_id, pe)
        cp_bal = await _balance_at_period_end(db, p.counterparty_tenant_id, p.counterparty_qbo_account_id, pe)
        if my_bal is None or cp_bal is None:
            # One side has no synced balance → can't eliminate; the side we DO
            # have still sits in the consolidation. Flag it.
            unmatched.append({
                "account_label": await _account_label_from_snapshot(db, tenant_id, p.my_qbo_account_id),
                "company_name": company_name(tenant_id),
                "my_balance": str((my_bal or Decimal("0")).quantize(Decimal("0.01"))),
                "counterparty_balance": (
                    str(cp_bal.quantize(Decimal("0.01"))) if cp_bal is not None else None
                ),
                "reason": "Counterparty balance not synced for this period",
            })
            continue
        diff = (my_bal + cp_bal).quantize(Decimal("0.01"))
        if abs(diff) > Decimal("1.00"):
            # Sides don't net → don't eliminate (would leave a residual); flag it.
            unmatched.append({
                "account_label": await _account_label_from_snapshot(db, tenant_id, p.my_qbo_account_id),
                "company_name": company_name(tenant_id),
                "my_balance": str(my_bal.quantize(Decimal("0.01"))),
                "counterparty_balance": str(cp_bal.quantize(Decimal("0.01"))),
                "reason": f"Sides don't net — off by ${abs(diff)}",
            })
            continue
        # eliminate full signed balance on each side so it nets to 0
        elim_by_acct[(tenant_id, p.my_qbo_account_id)] = -my_bal
        elim_by_acct[(p.counterparty_tenant_id, p.counterparty_qbo_account_id)] = -cp_bal

    # Pull latest snapshot per (tenant, qbo_account_id) AT OR BEFORE period_end
    snap_rows = (await db.execute(
        select(GlBalanceSnapshot)
        .where(
            GlBalanceSnapshot.tenant_id.in_(consol_tenants),
            GlBalanceSnapshot.period_end <= pe,
        ),
        execution_options={"skip_tenant_filter": True},
    )).scalars().all()
    latest: dict[tuple[uuid.UUID, str], GlBalanceSnapshot] = {}
    for s in snap_rows:
        key = (s.tenant_id, s.qbo_account_id)
        prior = latest.get(key)
        if prior is None or s.period_end > prior.period_end:
            latest[key] = s

    # Map QBO AccountType → high-level FS category for grouping
    def fs_cat(t: str) -> str:
        t = (t or "").lower()
        if t in ("bank", "accounts receivable", "other current asset", "fixed asset", "other asset"):
            return "Assets"
        if t in ("credit card", "accounts payable", "other current liability", "long term liability"):
            return "Liabilities"
        if t == "equity":
            return "Equity"
        if t in ("income", "other income"):
            return "Revenue"
        if t in ("expense", "cost of goods sold", "other expense"):
            return "Expenses"
        return "Other"

    rows_out: list[ConsolidatedRow] = []
    raw_by_cat: dict[str, Decimal] = {}
    elim_by_cat: dict[str, Decimal] = {}

    for (tid, qid), snap in latest.items():
        cat = fs_cat(snap.account_type)
        elim_amount = elim_by_acct.get((tid, qid), Decimal("0"))
        consol = (snap.balance + elim_amount).quantize(Decimal("0.01"))
        raw_by_cat[cat] = raw_by_cat.get(cat, Decimal("0")) + snap.balance
        elim_by_cat[cat] = elim_by_cat.get(cat, Decimal("0")) + elim_amount

        label = f"{(snap.account_number or '').strip()} {snap.account_name}".strip()
        rows_out.append(ConsolidatedRow(
            fs_category=cat,
            account_label=label,
            tenant_id=str(tid),
            company_name=company_name(tid),
            qbo_account_id=qid,
            raw_balance=str(snap.balance.quantize(Decimal("0.01"))),
            elimination=str(elim_amount.quantize(Decimal("0.01"))),
            consolidated=str(consol),
            is_eliminated_row=(elim_amount != Decimal("0")),
        ))

    # Sort: category order then company then account_label
    cat_order = {"Assets": 1, "Liabilities": 2, "Equity": 3, "Revenue": 4, "Expenses": 5, "Other": 9}
    rows_out.sort(key=lambda r: (
        cat_order.get(r.fs_category, 9), r.company_name.lower(), r.account_label.lower()
    ))

    companies_out = [
        {"tenant_id": str(tid), "name": company_name(tid)}
        for tid in consol_tenants
    ]
    companies_out.sort(key=lambda c: c["name"].lower())

    totals: dict = {}
    for cat in ("Assets", "Liabilities", "Equity", "Revenue", "Expenses"):
        raw = raw_by_cat.get(cat, Decimal("0")).quantize(Decimal("0.01"))
        elim = elim_by_cat.get(cat, Decimal("0")).quantize(Decimal("0.01"))
        totals[cat] = {
            "raw":          str(raw),
            "elimination":  str(elim),
            "consolidated": str((raw + elim).quantize(Decimal("0.01"))),
        }

    # Integrity check: a balanced consolidation nets to ~0 in debit-positive
    # terms (each entity's TB balances, and matched eliminations net to 0).
    consol_sum = sum(
        (Decimal(totals[c]["consolidated"])
         for c in ("Assets", "Liabilities", "Equity", "Revenue", "Expenses")),
        Decimal("0"),
    ).quantize(Decimal("0.01"))
    balanced = abs(consol_sum) <= Decimal("1.00")

    return ConsolidatedTbResponse(
        period_end=pe.isoformat(),
        companies=companies_out,
        rows=rows_out,
        totals=totals,
        balanced=balanced,
        imbalance=str(consol_sum),
        unmatched=unmatched,
    )


# ── Excel exports ──────────────────────────────────────────────────────────

@router.get("/eliminations.xlsx")
async def export_eliminations_xlsx(
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    period_end: str = Query(..., description="Period end YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
):
    """Excel export of the eliminations report."""
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    from modules.exports.xlsx_builder import (
        add_sheet_title,
        freeze_header,
        register_styles,
        set_column_widths,
        write_row,
        write_table_header,
    )

    data = await get_eliminations(tenant_id, user, period_end, db)

    wb = Workbook()
    register_styles(wb)
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Eliminations")
    ws.title = "Eliminations"

    hdr_row = add_sheet_title(ws, "Intercompany Eliminations",
                              subtitle=f"Period end: {data.period_end}")

    headers = ["Status", "My account", "My balance", "Counterparty", "Counterparty balance", "Diff"]
    write_table_header(ws, hdr_row, headers)
    set_column_widths(ws, [12, 38, 16, 42, 18, 16])

    r = hdr_row + 1
    for row in data.rows:
        status_label = {
            "matched": "Matched",
            "mismatch": "Mismatch",
            "one_side_missing": "Side missing",
        }.get(row.status, row.status)
        write_row(ws, r, [
            (status_label,            "nx_cell_text"),
            (row.my_account_label,    "nx_cell_text"),
            (Decimal(row.my_balance), "nx_cell_money"),
            (row.counterparty_label,  "nx_cell_text"),
            (Decimal(row.counterparty_balance), "nx_cell_money"),
            (Decimal(row.diff),       "nx_cell_money"),
        ])
        r += 1

    # Totals row
    write_row(ws, r, [
        ("Total to eliminate", "nx_total_label"),
        ("",                   "nx_total_label"),
        ("",                   "nx_total_label"),
        (f"{data.totals.get('matched_count', 0)} matched · {data.totals.get('mismatch_count', 0)} mismatch", "nx_total_label"),
        ("",                   "nx_total_label"),
        (Decimal(data.totals.get("total_to_eliminate", "0")), "nx_total_money"),
    ])

    freeze_header(ws, hdr_row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"intercompany_eliminations_{data.period_end}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/consolidated-tb.xlsx")
async def export_consolidated_tb_xlsx(
    tenant_id: CurrentTenantId,
    user: CurrentUser,
    period_end: str = Query(..., description="Period end YYYY-MM-DD"),
    db: AsyncSession = Depends(get_db),
):
    """Excel export of the consolidated TB with elimination column."""
    from io import BytesIO

    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook

    from modules.exports.xlsx_builder import (
        add_sheet_title,
        freeze_header,
        register_styles,
        set_column_widths,
        write_row,
        write_table_header,
    )

    data = await get_consolidated_tb(tenant_id, user, period_end, db)

    wb = Workbook()
    register_styles(wb)
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Consolidated TB")
    ws.title = "Consolidated TB"

    subtitle = (
        f"Period end: {data.period_end} · "
        f"Entities: {', '.join(c['name'] for c in data.companies)}"
    )
    hdr_row = add_sheet_title(ws, "Consolidated Trial Balance", subtitle=subtitle)

    headers = ["Category", "Company", "Account", "Raw balance", "Elimination", "Consolidated"]
    write_table_header(ws, hdr_row, headers)
    set_column_widths(ws, [14, 28, 36, 16, 16, 18])

    r = hdr_row + 1
    current_cat: str | None = None
    for row in data.rows:
        # Insert blank separator + category header on category change
        if row.fs_category != current_cat:
            current_cat = row.fs_category
            # blank row keeps the section break visible
            write_row(ws, r, [("", "nx_cell_text") for _ in range(6)])
            r += 1

        write_row(ws, r, [
            (row.fs_category,             "nx_cell_text"),
            (row.company_name,            "nx_cell_text"),
            (row.account_label,           "nx_cell_text"),
            (Decimal(row.raw_balance),    "nx_cell_money"),
            (Decimal(row.elimination),    "nx_cell_money"),
            (Decimal(row.consolidated),   "nx_cell_money"),
        ])
        r += 1

    # Category totals
    r += 1
    for cat in ("Assets", "Liabilities", "Equity", "Revenue", "Expenses"):
        t = data.totals.get(cat) or {}
        if not t:
            continue
        write_row(ws, r, [
            (f"Total {cat}",                 "nx_total_label"),
            ("",                              "nx_total_label"),
            ("",                              "nx_total_label"),
            (Decimal(t.get("raw", "0")),     "nx_total_money"),
            (Decimal(t.get("elimination", "0")), "nx_total_money"),
            (Decimal(t.get("consolidated", "0")), "nx_total_money"),
        ])
        r += 1

    freeze_header(ws, hdr_row)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"consolidated_tb_{data.period_end}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
