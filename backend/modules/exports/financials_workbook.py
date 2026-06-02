"""
Financial Package workbook (.xlsx) — the full GAAP close binder.

Builds a multi-sheet workbook from Nordavix synced data (with a QuickBooks
fallback for cash flow), or the whole thing live from QBO when source="quickbooks":

  Cover · Income Statement · Balance Sheet · Statement of Cash Flows ·
  Trial Balance · Cash & Cash Equivalents · AR Aging · AP Aging ·
  Prepaids · Fixed Assets & Depreciation · Accruals · Leases · Loans ·
  Equity Roll-forward · Reconciliation Summary

Two public entry points:
  - build_financials_workbook(...)        → the full package
  - build_single_financial_workbook(...)  → cover + one schedule

Reuses the monochrome styling + schedule sheet builders that already power the
Period Export, and the statement builders that power the Financial Package PDF.
No color anywhere (CPA-workpaper aesthetic), consistent with the rest of exports.
"""
from __future__ import annotations

import logging
import uuid
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from modules.exports.period_workbook import (
    _build_accruals_sheet,
    _build_fixed_assets_sheet,
    _build_leases_sheet,
    _build_loans_sheet,
    _build_prepaids_sheet,
    _build_recons_sheet,
    build_cover_sheet,
)
from modules.exports.xlsx_builder import (
    add_sheet_title,
    fmt_date,
    freeze_header,
    register_styles,
    safe_dec,
    set_column_widths,
    write_row,
    write_table_header,
)

logger = logging.getLogger(__name__)

# Monochrome tokens + money format (match xlsx_builder).
_DARK = "111827"
_RULE = "111827"
_SIDE_MED = Side(style="medium", color=_RULE)
_SIDE_DBL = Side(style="double", color=_RULE)
_MONEY_FMT = '_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-'

# Ordered sheet catalog — slug → display label. Drives the full-package order,
# the individual-export endpoint, and the frontend schedule list.
FINANCIAL_SHEETS: list[tuple[str, str]] = [
    ("income-statement",       "Income Statement"),
    ("balance-sheet",          "Balance Sheet"),
    ("cash-flow",              "Statement of Cash Flows"),
    ("trial-balance",          "Trial Balance"),
    ("cash",                   "Cash & Cash Equivalents"),
    ("ar-aging",               "Accounts Receivable Aging"),
    ("ap-aging",               "Accounts Payable Aging"),
    ("prepaids",               "Prepaid Expense Schedule"),
    ("fixed-assets",           "Fixed Assets & Depreciation"),
    ("accruals",               "Accrued Expense Schedule"),
    ("leases",                 "Lease Schedule"),
    ("loans",                  "Loan Schedule"),
    ("equity",                 "Equity Roll-forward"),
    ("reconciliation-summary", "Reconciliation Summary"),
]
FINANCIAL_SHEET_LABELS = dict(FINANCIAL_SHEETS)


# ── Build context (lazy-loads statements + snapshots, cached) ───────────────

@dataclass
class _Ctx:
    db: object
    tenant_id: uuid.UUID
    period_end: date
    period_start: date | None
    comparative: bool
    source: str
    _stmts: dict = field(default_factory=dict)
    _end_snap: list | None = None
    _beg_snap: tuple | None = None  # (rows, snapshot_date)

    @property
    def pe(self) -> date:
        return self.period_end

    @property
    def ps(self) -> date | None:
        return self.period_start

    async def statement(self, kind: str):
        if kind not in self._stmts:
            from modules.financials.router import _build_statement
            ps = self.period_start if kind in ("income_statement", "cash_flow") else None
            self._stmts[kind] = await _build_statement(
                self.tenant_id, self.db, self.period_end, kind, self.comparative,
                source=self.source, period_start=ps,
            )
        return self._stmts[kind]

    async def end_snapshot(self) -> list:
        if self._end_snap is None:
            from modules.financials.internal import _load_snapshot
            self._end_snap = await _load_snapshot(self.db, self.tenant_id, self.period_end)
        return self._end_snap

    async def begin_snapshot(self) -> tuple:
        if self._beg_snap is None:
            from datetime import timedelta

            from modules.financials.internal import load_snapshot_on_or_before
            start = self.period_start or date(self.period_end.year, 1, 1)
            self._beg_snap = await load_snapshot_on_or_before(
                self.db, self.tenant_id, start - timedelta(days=1))
        return self._beg_snap


def _money(s) -> float | None:
    if s is None or s == "":
        return None
    try:
        return float(s)
    except (TypeError, ValueError):
        return None


# ── Statement sheets (IS / BS / CFS) — render a StatementOut ──────────────────

def _stmt_value_cell(ws, r: int, col: int, raw, *, is_header: bool, is_total: bool, side) -> None:
    v = _money(raw)
    c = ws.cell(row=r, column=col, value=("" if (is_header or v is None) else v))
    c.font = Font(name="Calibri", size=10, bold=is_total, color=_DARK)
    c.alignment = Alignment(horizontal="right", vertical="center")
    if not is_header and v is not None:
        c.number_format = _MONEY_FMT
    if side:
        c.border = Border(top=side)


def _render_statement(ws, stmt) -> None:
    header_row = add_sheet_title(ws, stmt.company, subtitle=f"{stmt.title} — {stmt.subtitle}")
    has_comp = stmt.comparative_label is not None
    set_column_widths(ws, [56, 20] + ([20] if has_comp else []))
    write_table_header(ws, header_row, ["", stmt.period_label] + ([stmt.comparative_label] if has_comp else []))
    freeze_header(ws, header_row)

    r = header_row + 1
    for row in stmt.rows:
        is_header = row.kind == "section_header"
        is_total = row.kind in ("total", "subtotal", "computed", "grand_total")
        side = _SIDE_DBL if row.kind == "grand_total" else (_SIDE_MED if is_total else None)
        label = ("    " * max(row.level, 0)) + row.label

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name="Calibri", size=10, bold=(is_total or is_header), color=_DARK)
        lc.alignment = Alignment(horizontal="left", vertical="center")
        if side:
            lc.border = Border(top=side)

        _stmt_value_cell(ws, r, 2, row.current, is_header=is_header, is_total=is_total, side=side)
        if has_comp:
            _stmt_value_cell(ws, r, 3, row.prior, is_header=is_header, is_total=is_total, side=side)
        r += 1

    # Source + notes footer
    notes = list(getattr(stmt, "notes", []) or [])
    if notes:
        r += 1
        for n in notes:
            cell = ws.cell(row=r, column=1, value=f"Note: {n}")
            cell.font = Font(name="Calibri", size=9, italic=True, color="6B7280")
            r += 1


async def _sheet_income_statement(wb: Workbook, ctx: _Ctx) -> None:
    _render_statement(wb.create_sheet("Income Statement"), await ctx.statement("income_statement"))


async def _sheet_balance_sheet(wb: Workbook, ctx: _Ctx) -> None:
    _render_statement(wb.create_sheet("Balance Sheet"), await ctx.statement("balance_sheet"))


async def _sheet_cash_flow(wb: Workbook, ctx: _Ctx) -> None:
    _render_statement(wb.create_sheet("Cash Flows"), await ctx.statement("cash_flow"))


# ── Trial Balance ────────────────────────────────────────────────────────────

async def _sheet_trial_balance(wb: Workbook, ctx: _Ctx) -> None:
    ws = wb.create_sheet("Trial Balance")
    rows = await ctx.end_snapshot()
    header_row = add_sheet_title(ws, "Trial Balance", subtitle=f"As of {fmt_date(ctx.pe)}")
    set_column_widths(ws, [12, 46, 24, 18, 18])
    write_table_header(ws, header_row, ["Acct #", "Account", "Type", "Debit", "Credit"])
    freeze_header(ws, header_row)

    if not rows:
        write_row(ws, header_row + 1, [
            ("No GL snapshot for this period — sync first.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(4)],
        ])
        return

    r = header_row + 1
    tot_d = tot_c = Decimal("0")
    for a in sorted(rows, key=lambda x: (x.account_number or "", x.account_name)):
        bal = a.balance
        deb = bal if bal > 0 else Decimal("0")
        cre = -bal if bal < 0 else Decimal("0")
        tot_d += deb
        tot_c += cre
        write_row(ws, r, [
            (a.account_number or "", "nx_cell_text"),
            (a.account_name, "nx_cell_text"),
            (a.account_type, "nx_cell_text"),
            (deb if deb else "", "nx_cell_money"),
            (cre if cre else "", "nx_cell_money"),
        ])
        r += 1
    write_row(ws, r, [
        ("", "nx_total_label"), ("TOTAL", "nx_total_label"), ("", "nx_total_label"),
        (tot_d, "nx_total_money"), (tot_c, "nx_total_money"),
    ])


# ── Cash & Cash Equivalents ──────────────────────────────────────────────────

async def _sheet_cash(wb: Workbook, ctx: _Ctx) -> None:
    ws = wb.create_sheet("Cash")
    rows = await ctx.end_snapshot()
    banks = [a for a in rows if a.account_type == "Bank"]
    header_row = add_sheet_title(ws, "Cash and Cash Equivalents", subtitle=f"As of {fmt_date(ctx.pe)}")
    set_column_widths(ws, [12, 50, 20])
    write_table_header(ws, header_row, ["Acct #", "Account", "Balance"])
    freeze_header(ws, header_row)

    if not banks:
        write_row(ws, header_row + 1, [
            ("No bank accounts in snapshot.", "nx_cell_muted"), ("", "nx_cell_text"), ("", "nx_cell_text"),
        ])
        return

    r = header_row + 1
    tot = Decimal("0")
    for a in banks:
        tot += a.balance
        write_row(ws, r, [
            (a.account_number or "", "nx_cell_text"),
            (a.account_name, "nx_cell_text"),
            (a.balance, "nx_cell_money"),
        ])
        r += 1
    write_row(ws, r, [("", "nx_total_label"), ("TOTAL CASH", "nx_total_label"), (tot, "nx_total_money")])


# ── AR / AP aging (QBO best-effort; falls back to GL balance) ─────────────────

async def _qbo_conn(ctx: _Ctx):
    from sqlalchemy import select

    from models.qbo_connection import QboConnection
    return (await ctx.db.execute(
        select(QboConnection).where(QboConnection.tenant_id == ctx.tenant_id),
        execution_options={"skip_tenant_filter": True},
    )).scalar_one_or_none()


async def _aging_sheet(wb: Workbook, ctx: _Ctx, *, title: str, sheet: str,
                       entity_header: str, fetch, gl_type: str) -> None:
    ws = wb.create_sheet(sheet)
    header_row = add_sheet_title(ws, title, subtitle=f"As of {fmt_date(ctx.pe)}")
    headers = [entity_header, "Current", "1–30", "31–60", "61–90", "Over 90", "Total"]
    set_column_widths(ws, [36, 14, 14, 14, 14, 14, 16])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    rows: list[dict] = []
    try:
        conn = await _qbo_conn(ctx)
        if conn is not None:
            rows, _src = await fetch(conn, ctx.db, ctx.pe)
    except Exception:
        logger.exception("%s: aging fetch failed", sheet)
        rows = []

    if not rows:
        end_rows = await ctx.end_snapshot()
        gl = sum((a.balance for a in end_rows if a.account_type == gl_type), Decimal("0"))
        write_row(ws, header_row + 1, [
            ("Aging detail unavailable (connect QuickBooks) — GL balance shown.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(5)],
            (abs(gl), "nx_cell_money"),
        ])
        return

    r = header_row + 1
    cols = ("current", "1_30", "31_60", "61_90", "over_90", "total")
    totals = dict.fromkeys(cols, Decimal("0"))
    for it in rows:
        vals = {k: safe_dec(it.get(k)) for k in cols}
        for k in cols:
            totals[k] += vals[k]
        write_row(ws, r, [
            (it.get("label") or "", "nx_cell_text"),
            *[(vals[k], "nx_cell_money") for k in cols],
        ])
        r += 1
    write_row(ws, r, [
        ("TOTAL", "nx_total_label"),
        *[(totals[k], "nx_total_money") for k in cols],
    ])


async def _sheet_ar_aging(wb: Workbook, ctx: _Ctx) -> None:
    from modules.recons.overview import _ar_subledger_rows
    await _aging_sheet(wb, ctx, title="Accounts Receivable Aging", sheet="AR Aging",
                       entity_header="Customer", fetch=_ar_subledger_rows,
                       gl_type="Accounts Receivable")


async def _sheet_ap_aging(wb: Workbook, ctx: _Ctx) -> None:
    from modules.recons.overview import _ap_subledger_rows
    await _aging_sheet(wb, ctx, title="Accounts Payable Aging", sheet="AP Aging",
                       entity_header="Vendor", fetch=_ap_subledger_rows,
                       gl_type="Accounts Payable")


# ── Equity roll-forward ──────────────────────────────────────────────────────

async def _sheet_equity(wb: Workbook, ctx: _Ctx) -> None:
    from modules.financials.internal import (
        _EQUITY_TYPES,
        _compute_net_income,
        _presented_sum,
    )
    ws = wb.create_sheet("Equity Roll-forward")
    end_rows = await ctx.end_snapshot()
    beg_rows, beg_date = await ctx.begin_snapshot()
    header_row = add_sheet_title(ws, "Statement of Stockholders' Equity",
                                 subtitle=f"For the period ending {fmt_date(ctx.pe)}")
    set_column_widths(ws, [56, 20])
    write_table_header(ws, header_row, ["", "Amount"])
    freeze_header(ws, header_row)

    if not end_rows:
        write_row(ws, header_row + 1, [
            ("No GL snapshot for this period — sync first.", "nx_cell_muted"), ("", "nx_cell_text"),
        ])
        return

    end_eq = _presented_sum(end_rows, _EQUITY_TYPES)
    end_ni = _compute_net_income(end_rows)

    if beg_rows and beg_date is not None:
        beg_eq = _presented_sum(beg_rows, _EQUITY_TYPES)
        beg_ni = _compute_net_income(beg_rows)
        same_year = beg_date.year == ctx.pe.year
        period_ni = (end_ni - beg_ni) if same_year else end_ni
        beginning_total = beg_eq + beg_ni
        ending_total = end_eq + end_ni
        contributions = ending_total - beginning_total - period_ni
        lines = [
            ("Total equity, beginning of period", beginning_total, False),
            ("Net income for the period", period_ni, False),
            ("Owner contributions / (distributions), net", contributions, False),
            ("Total equity, end of period", ending_total, True),
        ]
    else:
        lines = [
            ("Contributed capital and retained earnings", end_eq, False),
            ("Current-year net income", end_ni, False),
            ("Total equity, end of period", end_eq + end_ni, True),
        ]

    r = header_row + 1
    for label, amt, is_total in lines:
        write_row(ws, r, [
            (label, "nx_cell_text"),
            (amt, "nx_total_money" if is_total else "nx_cell_money"),
        ])
        r += 1


# ── Schedule sheets reused verbatim from period_workbook ─────────────────────

async def _sheet_prepaids(wb, ctx):       await _build_prepaids_sheet(wb, ctx.db, ctx.tenant_id, ctx.pe)
async def _sheet_fixed_assets(wb, ctx):   await _build_fixed_assets_sheet(wb, ctx.db, ctx.tenant_id, ctx.pe)
async def _sheet_accruals(wb, ctx):       await _build_accruals_sheet(wb, ctx.db, ctx.tenant_id, ctx.pe)
async def _sheet_leases(wb, ctx):         await _build_leases_sheet(wb, ctx.db, ctx.tenant_id, ctx.pe)
async def _sheet_loans(wb, ctx):          await _build_loans_sheet(wb, ctx.db, ctx.tenant_id, ctx.pe)
async def _sheet_recon_summary(wb, ctx):  await _build_recons_sheet(wb, ctx.db, ctx.tenant_id, ctx.pe)


# ── Dispatch ─────────────────────────────────────────────────────────────────

_BUILDERS = {
    "income-statement":       _sheet_income_statement,
    "balance-sheet":          _sheet_balance_sheet,
    "cash-flow":              _sheet_cash_flow,
    "trial-balance":          _sheet_trial_balance,
    "cash":                   _sheet_cash,
    "ar-aging":               _sheet_ar_aging,
    "ap-aging":               _sheet_ap_aging,
    "prepaids":               _sheet_prepaids,
    "fixed-assets":           _sheet_fixed_assets,
    "accruals":               _sheet_accruals,
    "leases":                 _sheet_leases,
    "loans":                  _sheet_loans,
    "equity":                 _sheet_equity,
    "reconciliation-summary": _sheet_recon_summary,
}


def _error_sheet(wb: Workbook, label: str, exc: Exception) -> None:
    ws = wb.create_sheet(label[:31])
    ws.column_dimensions["A"].width = 64
    ws["A1"] = label
    ws["A1"].style = "nx_h2"
    ws["A3"] = "Could not build this section."
    ws["A3"].style = "nx_label"
    ws["A4"] = f"Reason: {type(exc).__name__}: {str(exc)[:200]}"
    ws["A4"].style = "nx_cell_muted"


def _period_label(period_end: date, period_start: date | None) -> str:
    if period_start is not None:
        return f"Period {fmt_date(period_start)} – {fmt_date(period_end)}"
    return f"Period ending {fmt_date(period_end)}"


def _new_wb() -> Workbook:
    wb = Workbook()
    register_styles(wb)
    if wb.active is not None:
        wb.remove(wb.active)
    return wb


def _save(wb: Workbook) -> bytes:
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


async def build_financials_workbook(
    *,
    db,
    tenant_id: uuid.UUID,
    period_end: date,
    period_start: date | None,
    comparative: bool,
    source: str,
    company_name: str,
    generated_by_name: str,
) -> bytes:
    """Full GAAP close binder. Each sheet is isolated — one failure becomes a
    placeholder sheet so the rest of the workbook still downloads."""
    ctx = _Ctx(db=db, tenant_id=tenant_id, period_end=period_end,
               period_start=period_start, comparative=comparative, source=source)
    wb = _new_wb()
    build_cover_sheet(
        wb,
        company_name=company_name,
        package_title="Financial statement package",
        period_label=_period_label(period_end, period_start),
        generated_by=generated_by_name,
        contents_text="Income Statement, Balance Sheet, Cash Flows, Trial Balance, "
                      "Cash, AR/AP aging, Prepaids, Fixed Assets, Accruals, Leases, "
                      "Loans, Equity, Reconciliation summary",
        footer_text=(
            "Built from Nordavix synced data (GL snapshots, schedules, and "
            "reconciliations). Cash flow is indirect-method, derived from the "
            "beginning and ending balance-sheet positions."
        ),
    )
    for slug, label in FINANCIAL_SHEETS:
        try:
            await _BUILDERS[slug](wb, ctx)
        except Exception as exc:
            logger.exception("Financials workbook: %s sheet failed", slug)
            _error_sheet(wb, label, exc)
    return _save(wb)


async def build_single_financial_workbook(
    *,
    slug: str,
    db,
    tenant_id: uuid.UUID,
    period_end: date,
    period_start: date | None,
    comparative: bool,
    source: str,
    company_name: str,
    generated_by_name: str,
) -> bytes:
    """Cover + a single financial schedule."""
    if slug not in _BUILDERS:
        raise ValueError(f"Unknown financial schedule: {slug}")
    label = FINANCIAL_SHEET_LABELS[slug]
    ctx = _Ctx(db=db, tenant_id=tenant_id, period_end=period_end,
               period_start=period_start, comparative=comparative, source=source)
    wb = _new_wb()
    build_cover_sheet(
        wb,
        company_name=company_name,
        package_title=label,
        period_label=_period_label(period_end, period_start),
        generated_by=generated_by_name,
        contents_text=label,
        footer_text="Generated from Nordavix synced data.",
    )
    try:
        await _BUILDERS[slug](wb, ctx)
    except Exception as exc:
        logger.exception("Financials single export: %s failed", slug)
        _error_sheet(wb, label, exc)
    return _save(wb)


def filename_for(slug: str, company_name: str, period_end: date) -> str:
    import re
    safe = re.sub(r"[^A-Za-z0-9 _-]+", "", company_name or "Nordavix").strip()
    safe = re.sub(r"\s+", "_", safe)[:50] or "Nordavix"
    return f"{safe}_{slug}_{period_end.isoformat()}.xlsx"
