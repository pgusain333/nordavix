"""
Build the Period Export workbook (.xlsx) for one (tenant, period_end).

Produces a multi-sheet workbook in memory (BytesIO) — caller streams it
back via StreamingResponse. Each sheet is built by a small function;
no shared state, so adding/removing sheets is trivial.

Data sources:
  - Reconciliations sheet      → modules.recons.overview snapshot reader
  - Prepaid / Accrual / etc.   → schedule_* models direct query
  - Audit log                  → AuditLog model, last 90 days
"""
from __future__ import annotations

import logging
import uuid
from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from models.audit_log import AuditLog
from models.schedule import (
    ScheduleAccrual,
    ScheduleFixedAsset,
    ScheduleLease,
    ScheduleLoan,
    SchedulePrepaid,
)
from modules.exports.xlsx_builder import (
    fmt_date,
    freeze_header,
    register_styles,
    safe_dec,
    set_column_widths,
    write_row,
    write_table_header,
)
from modules.schedules import calc

logger = logging.getLogger(__name__)


# ── Main entry point ──────────────────────────────────────────────────

async def build_period_workbook(
    *,
    db: AsyncSession,
    tenant_id: uuid.UUID,
    period_end: date,
    company_name: str,
    generated_by_name: str,
) -> bytes:
    """Build the workbook + return the raw bytes ready to stream.

    Each section is wrapped in its own try/except so one failed sheet
    (e.g. a schedule with a stale column, an audit-log permission
    issue) doesn't 500 the whole download. Failed sheets show a
    one-line "could not load" placeholder + the error class name; the
    full traceback is logged server-side so we can debug without
    blocking the user.
    """
    wb = Workbook()
    register_styles(wb)

    # Remove the default sheet — we'll add named ones below
    if wb.active is not None:
        wb.remove(wb.active)

    _build_cover_sheet(wb, company_name=company_name, period_end=period_end,
                       generated_by=generated_by_name)

    sections = [
        ("Reconciliations", _build_recons_sheet),
        ("Prepaids",        _build_prepaids_sheet),
        ("Accruals",        _build_accruals_sheet),
        ("Fixed Assets",    _build_fixed_assets_sheet),
        ("Leases",          _build_leases_sheet),
        ("Loans",           _build_loans_sheet),
        ("Audit Log",       _build_audit_log_sheet),
    ]
    for label, builder in sections:
        try:
            await builder(wb, db, tenant_id, period_end)
        except Exception as exc:
            logger.exception("Period export: %s sheet build failed", label)
            _build_error_sheet(wb, label, exc)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _build_error_sheet(wb: Workbook, label: str, exc: Exception) -> None:
    """Drop a placeholder sheet so the user knows something failed
    but the rest of the workbook still downloads. Keeps the sheet
    order intact so saved-file references don't shift."""
    ws = wb.create_sheet(label)
    ws.column_dimensions["A"].width = 60
    ws["A1"] = label
    ws["A1"].style = "nx_h2"
    ws["A3"] = "Could not build this section."
    ws["A3"].style = "nx_label"
    ws["A4"] = f"Reason: {type(exc).__name__}: {str(exc)[:200]}"
    ws["A4"].style = "nx_cell_muted"
    ws["A6"] = (
        "The rest of the workbook downloaded normally. If this keeps "
        "happening, share this filename + the period with support; the "
        "full traceback is in the server logs."
    )
    ws["A6"].style = "nx_cell_muted"


# ── Sheet 1: Cover ─────────────────────────────────────────────────────

def build_cover_sheet(
    wb: Workbook,
    *,
    company_name: str,
    package_title: str,
    period_label: str,
    generated_by: str,
    contents_text: str,
    footer_text: str | None = None,
) -> None:
    """Reusable cover sheet. Used by Period, Flux, Recon, Schedules exports.

    Args:
      package_title: Large H1 line under the wordmark (e.g. "Month-end close package").
      period_label:  Human period string (e.g. "Period ending 04-30-2026" or "FY2026 YTD").
      contents_text: One-line description of what's in this workbook.
      footer_text:   Optional italic footer note.
    """
    ws = wb.create_sheet("Cover", 0)

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 60

    ws["A2"] = "NORDAVIX"
    ws["A2"].font = ws["A2"].font.copy(name="Calibri", size=10, bold=True, color="111827")

    ws["A4"] = package_title
    ws["A4"].style = "nx_h1"
    ws.row_dimensions[4].height = 40

    ws["A6"] = company_name or "(workspace)"
    ws["A6"].font = ws["A6"].font.copy(name="Calibri", size=20, bold=True)

    rows = [
        ("PERIOD",        period_label),
        ("GENERATED",     fmt_date(datetime.utcnow().date()) + " · " + datetime.utcnow().strftime("%I:%M %p UTC")),
        ("GENERATED BY",  generated_by or "(unknown user)"),
        ("FILE CONTENTS", contents_text),
    ]
    for i, (label, value) in enumerate(rows, start=9):
        ws.cell(row=i, column=1, value=label).style = "nx_label"
        ws.cell(row=i, column=2, value=value).style = "nx_value"
        ws.row_dimensions[i].height = 22

    if footer_text:
        ws["A16"] = footer_text
        ws["A16"].font = ws["A16"].font.copy(name="Calibri", size=10, italic=True, color="6B7280")
        ws["A16"].alignment = ws["A16"].alignment.copy(wrap_text=True, vertical="top")
        ws.merge_cells("A16:B17")


def _build_cover_sheet(
    wb: Workbook, *, company_name: str, period_end: date, generated_by: str,
) -> None:
    """Period Export cover (thin wrapper around the reusable builder)."""
    build_cover_sheet(
        wb,
        company_name=company_name,
        package_title="Month-end close package",
        period_label=f"Period ending {fmt_date(period_end)}",
        generated_by=generated_by,
        contents_text="Reconciliations, Prepaids, Accruals, Fixed Assets, Leases, Loans, Audit log",
        footer_text=(
            "This workbook is generated directly from Nordavix's snapshots — "
            "the same source of truth that powers the in-app reconciliation "
            "dashboard and the executive PDF package."
        ),
    )


# ── Sheet 2: Reconciliations summary ───────────────────────────────────

async def _build_recons_sheet(
    wb: Workbook, db: AsyncSession, tenant_id: uuid.UUID, period_end: date,
) -> None:
    ws = wb.create_sheet("Reconciliations")

    # Read from the same snapshot machinery the dashboard uses
    try:
        from modules.recons.overview import read_overview_from_snapshots
        overview = await read_overview_from_snapshots(db, period_end)
    except Exception:
        logger.exception("Recons sheet: snapshot read failed")
        overview = None

    header_row = _title(ws, "Reconciliations", subtitle=f"As of {fmt_date(period_end)}")
    headers = [
        "Acct #", "Account", "Type", "GL Balance",
        "Subledger", "Variance", "Status", "Reviewer notes",
    ]
    set_column_widths(ws, [10, 38, 22, 16, 16, 16, 14, 40])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    if not overview:
        write_row(ws, header_row + 1, [
            ("No committed snapshot for this period.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(7)],
        ])
        return

    accounts = overview.get("accounts", []) or []
    if not accounts:
        write_row(ws, header_row + 1, [
            ("No balance-sheet accounts in this snapshot.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(7)],
        ])
        return

    sum_gl  = Decimal("0")
    sum_sl  = Decimal("0")
    sum_var = Decimal("0")

    row = header_row + 1
    for a in accounts:
        gl  = safe_dec(a.get("gl_balance"))
        sl  = safe_dec(a.get("subledger_balance"))
        var = safe_dec(a.get("variance"))
        sum_gl += gl
        sum_sl += sl
        sum_var += var

        write_row(ws, row, [
            (a.get("account_number") or "", "nx_cell_text"),
            (a.get("account_name") or "",   "nx_cell_text"),
            (a.get("account_type") or a.get("group_label") or "", "nx_cell_text"),
            (gl,  "nx_cell_money"),
            (sl,  "nx_cell_money"),
            (var, "nx_cell_money"),
            ((a.get("review_status") or "pending").title(), "nx_cell_text"),
            (a.get("review_notes") or "", "nx_cell_muted"),
        ])
        row += 1

    # Totals row
    write_row(ws, row, [
        ("", "nx_total_label"),
        ("TOTAL", "nx_total_label"),
        ("", "nx_total_label"),
        (sum_gl,  "nx_total_money"),
        (sum_sl,  "nx_total_money"),
        (sum_var, "nx_total_money"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
    ])


# ── Sheet 3: Prepaids ──────────────────────────────────────────────────

async def _build_prepaids_sheet(
    wb: Workbook, db: AsyncSession, tenant_id: uuid.UUID, period_end: date,
) -> None:
    ws = wb.create_sheet("Prepaids")

    items = list((await db.execute(
        select(SchedulePrepaid).where(SchedulePrepaid.is_active == True)  # noqa: E712
        .order_by(SchedulePrepaid.start_date.desc())
    )).scalars().all())

    header_row = _title(ws, "Prepaid Expense Schedule",
                        subtitle=f"Active items as of {fmt_date(period_end)}")
    headers = [
        "Vendor", "Description", "Invoice", "Start", "End",
        "Method", "Total", "Period Amortization",
        "Amortized to date", "Unamortized",
    ]
    set_column_widths(ws, [24, 36, 16, 12, 12, 14, 14, 18, 18, 16])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    if not items:
        write_row(ws, header_row + 1, [
            ("No active prepaid items.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(9)],
        ])
        return

    p_start, p_end = calc._period_bounds(period_end)
    sum_total = sum_per = sum_to_date = sum_unam = Decimal("0")
    row = header_row + 1
    for it in items:
        period_amort = calc._prepaid_period_expense(it, p_start, p_end)
        amort_to_date = calc._prepaid_amortized_through(it, p_end)
        unamortized = calc._prepaid_unamortized_as_of(it, p_end)
        sum_total    += Decimal(it.total_amount)
        sum_per      += period_amort
        sum_to_date  += amort_to_date
        sum_unam     += unamortized

        method = "Straight-line" if calc._prepaid_method(it) == "straight_line" else "Daily-rate"

        write_row(ws, row, [
            (it.vendor or "",       "nx_cell_text"),
            (it.description or "",  "nx_cell_text"),
            (it.reference or "",    "nx_cell_text"),
            (it.start_date,         "nx_cell_date"),
            (it.end_date,           "nx_cell_date"),
            (method,                "nx_cell_text"),
            (Decimal(it.total_amount), "nx_cell_money"),
            (period_amort,             "nx_cell_money"),
            (amort_to_date,            "nx_cell_money"),
            (unamortized,              "nx_cell_money"),
        ])
        row += 1

    write_row(ws, row, [
        ("", "nx_total_label"),
        ("TOTAL", "nx_total_label"),
        *[("", "nx_total_label") for _ in range(4)],
        (sum_total,    "nx_total_money"),
        (sum_per,      "nx_total_money"),
        (sum_to_date,  "nx_total_money"),
        (sum_unam,     "nx_total_money"),
    ])


# ── Sheet 4: Accruals ──────────────────────────────────────────────────

async def _build_accruals_sheet(
    wb: Workbook, db: AsyncSession, tenant_id: uuid.UUID, period_end: date,
) -> None:
    ws = wb.create_sheet("Accruals")

    items = list((await db.execute(
        select(ScheduleAccrual).where(ScheduleAccrual.is_active == True)  # noqa: E712
        .order_by(ScheduleAccrual.accrual_date.desc())
    )).scalars().all())

    header_row = _title(ws, "Accrued Expense Schedule",
                        subtitle=f"Active items as of {fmt_date(period_end)}")
    headers = [
        "Vendor", "Description", "Reference",
        "Accrual date", "Amount", "Reverses on", "Status",
    ]
    set_column_widths(ws, [26, 36, 16, 14, 14, 14, 14])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    if not items:
        write_row(ws, header_row + 1, [
            ("No active accruals.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(6)],
        ])
        return

    sum_amt = Decimal("0")
    row = header_row + 1
    for it in items:
        amt = Decimal(it.amount)
        sum_amt += amt
        status = "Reversed" if it.is_reversed else ("Pending reversal" if it.reverses_on else "Active")
        write_row(ws, row, [
            (it.vendor or "",       "nx_cell_text"),
            (it.description or "",  "nx_cell_text"),
            (it.reference or "",    "nx_cell_text"),
            (it.accrual_date,       "nx_cell_date"),
            (amt,                   "nx_cell_money"),
            (it.reverses_on if it.reverses_on else "", "nx_cell_date" if it.reverses_on else "nx_cell_text"),
            (status,                "nx_cell_text"),
        ])
        row += 1

    write_row(ws, row, [
        ("", "nx_total_label"),
        ("TOTAL", "nx_total_label"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        (sum_amt, "nx_total_money"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
    ])


# ── Sheet 5: Fixed Assets ──────────────────────────────────────────────

async def _build_fixed_assets_sheet(
    wb: Workbook, db: AsyncSession, tenant_id: uuid.UUID, period_end: date,
) -> None:
    ws = wb.create_sheet("Fixed Assets")

    items = list((await db.execute(
        select(ScheduleFixedAsset).where(ScheduleFixedAsset.is_active == True)  # noqa: E712
        .order_by(ScheduleFixedAsset.in_service_date.desc())
    )).scalars().all())

    header_row = _title(ws, "Fixed Asset Register",
                        subtitle=f"As of {fmt_date(period_end)}")
    headers = [
        "Description", "Category", "In service", "Cost",
        "Salvage", "Useful life (mo)", "Method",
        "Accum depreciation", "Net book value", "Disposed on",
    ]
    set_column_widths(ws, [36, 18, 12, 14, 14, 14, 14, 18, 16, 12])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    if not items:
        write_row(ws, header_row + 1, [
            ("No active fixed assets.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(9)],
        ])
        return

    sum_cost = sum_acc = sum_nbv = Decimal("0")
    row = header_row + 1
    for it in items:
        cost = Decimal(it.cost)
        salv = Decimal(it.salvage_value or 0)
        # Linear straight-line accumulated depreciation through period_end
        months_elapsed = calc._months_between(it.in_service_date, period_end)
        months_elapsed = min(months_elapsed, it.useful_life_months)
        depreciable = max(cost - salv, Decimal("0"))
        per_month = depreciable / Decimal(max(it.useful_life_months, 1))
        accum = (per_month * Decimal(months_elapsed)).quantize(Decimal("0.01"))
        nbv = cost - accum
        if it.disposed_on and it.disposed_on <= period_end:
            nbv = Decimal("0")
        sum_cost += cost
        sum_acc  += accum
        sum_nbv  += nbv

        write_row(ws, row, [
            (it.description or "",  "nx_cell_text"),
            (it.category or "",     "nx_cell_text"),
            (it.in_service_date,    "nx_cell_date"),
            (cost,                  "nx_cell_money"),
            (salv,                  "nx_cell_money"),
            (it.useful_life_months, "nx_cell_int"),
            ("Straight-line",       "nx_cell_text"),
            (accum,                 "nx_cell_money"),
            (nbv,                   "nx_cell_money"),
            (it.disposed_on if it.disposed_on else "", "nx_cell_date" if it.disposed_on else "nx_cell_text"),
        ])
        row += 1

    write_row(ws, row, [
        ("", "nx_total_label"),
        ("TOTAL", "nx_total_label"),
        ("", "nx_total_label"),
        (sum_cost, "nx_total_money"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        (sum_acc, "nx_total_money"),
        (sum_nbv, "nx_total_money"),
        ("", "nx_total_label"),
    ])


# ── Sheet 6: Leases ────────────────────────────────────────────────────

async def _build_leases_sheet(
    wb: Workbook, db: AsyncSession, tenant_id: uuid.UUID, period_end: date,
) -> None:
    ws = wb.create_sheet("Leases")

    items = list((await db.execute(
        select(ScheduleLease).where(ScheduleLease.is_active == True)  # noqa: E712
        .order_by(ScheduleLease.lease_start.desc())
    )).scalars().all())

    header_row = _title(ws, "Lease Schedule",
                        subtitle=f"Active leases as of {fmt_date(period_end)}")
    headers = [
        "Lessor", "Description", "Start", "End",
        "Monthly payment", "Discount rate %",
        "Initial ROU asset", "Initial liability", "Reference",
    ]
    set_column_widths(ws, [22, 30, 12, 12, 16, 14, 18, 18, 18])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    if not items:
        write_row(ws, header_row + 1, [
            ("No active leases.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(8)],
        ])
        return

    sum_pay = sum_rou = sum_liab = Decimal("0")
    row = header_row + 1
    for it in items:
        pay = Decimal(it.monthly_payment)
        rou = Decimal(it.initial_rou_asset or 0)
        liab = Decimal(it.initial_liability or 0)
        sum_pay  += pay
        sum_rou  += rou
        sum_liab += liab
        write_row(ws, row, [
            (it.lessor or "",      "nx_cell_text"),
            (it.description or "", "nx_cell_text"),
            (it.lease_start,       "nx_cell_date"),
            (it.lease_end,         "nx_cell_date"),
            (pay,                  "nx_cell_money"),
            (float(it.discount_rate_pct) if it.discount_rate_pct is not None else "", "nx_cell_text"),
            (rou,                  "nx_cell_money"),
            (liab,                 "nx_cell_money"),
            (it.reference or "",   "nx_cell_text"),
        ])
        row += 1

    write_row(ws, row, [
        ("", "nx_total_label"),
        ("TOTAL", "nx_total_label"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        (sum_pay,  "nx_total_money"),
        ("", "nx_total_label"),
        (sum_rou,  "nx_total_money"),
        (sum_liab, "nx_total_money"),
        ("", "nx_total_label"),
    ])


# ── Sheet 7: Loans ─────────────────────────────────────────────────────

async def _build_loans_sheet(
    wb: Workbook, db: AsyncSession, tenant_id: uuid.UUID, period_end: date,
) -> None:
    ws = wb.create_sheet("Loans")

    items = list((await db.execute(
        select(ScheduleLoan).where(ScheduleLoan.is_active == True)  # noqa: E712
        .order_by(ScheduleLoan.loan_date.desc())
    )).scalars().all())

    header_row = _title(ws, "Loan Schedule",
                        subtitle=f"Active loans as of {fmt_date(period_end)}")
    headers = [
        "Lender", "Description", "Loan date", "Original principal",
        "Rate %", "Term (mo)", "Payment type", "Monthly payment", "Reference",
    ]
    set_column_widths(ws, [22, 30, 12, 18, 10, 12, 14, 16, 18])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    if not items:
        write_row(ws, header_row + 1, [
            ("No active loans.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(8)],
        ])
        return

    sum_prin = sum_pay = Decimal("0")
    row = header_row + 1
    for it in items:
        prin = Decimal(it.original_principal)
        pay  = Decimal(it.monthly_payment or 0)
        sum_prin += prin
        sum_pay  += pay
        write_row(ws, row, [
            (it.lender or "",                 "nx_cell_text"),
            (it.description or "",            "nx_cell_text"),
            (it.loan_date,                    "nx_cell_date"),
            (prin,                            "nx_cell_money"),
            (float(it.interest_rate_pct or 0),"nx_cell_text"),
            (it.term_months,                  "nx_cell_int"),
            ((it.payment_type or "amortizing").replace("_", " ").title(), "nx_cell_text"),
            (pay,                             "nx_cell_money"),
            (it.reference or "",              "nx_cell_text"),
        ])
        row += 1

    write_row(ws, row, [
        ("", "nx_total_label"),
        ("TOTAL", "nx_total_label"),
        ("", "nx_total_label"),
        (sum_prin, "nx_total_money"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        (sum_pay,  "nx_total_money"),
        ("", "nx_total_label"),
    ])


# ── Sheet 8: Audit Log ─────────────────────────────────────────────────

async def _build_audit_log_sheet(
    wb: Workbook, db: AsyncSession, tenant_id: uuid.UUID, period_end: date,
) -> None:
    ws = wb.create_sheet("Audit Log")

    since = datetime.combine(period_end, datetime.min.time()) - timedelta(days=90)
    rows = list((await db.execute(
        select(AuditLog)
        .where(AuditLog.created_at >= since)
        .order_by(desc(AuditLog.created_at))
        .limit(2000)
    )).scalars().all())

    header_row = _title(ws, "Audit Log",
                        subtitle=f"Last 90 days ending {fmt_date(period_end)}")
    headers = ["Timestamp (UTC)", "Action", "Summary", "User"]
    set_column_widths(ws, [22, 28, 60, 22])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    if not rows:
        write_row(ws, header_row + 1, [
            ("No events in window.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(3)],
        ])
        return

    row = header_row + 1
    for r in rows:
        write_row(ws, row, [
            (r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "", "nx_cell_text"),
            (r.action or "",  "nx_cell_text"),
            # AuditLog has no `summary` column — it's derived from event_data
            # (mirrors the /api/audit serializer). The old `r.summary` raised
            # AttributeError on the first row, so the section's try/except left
            # the sheet headers-only ("blank, no details").
            ((r.event_data or {}).get("summary") or r.action or "", "nx_cell_text"),
            (str(r.user_id) if r.user_id else "", "nx_cell_muted"),
        ])
        row += 1


# ── Small helper ──────────────────────────────────────────────────────

def _title(ws, title: str, *, subtitle: str | None = None) -> int:
    from modules.exports.xlsx_builder import add_sheet_title
    return add_sheet_title(ws, title, subtitle=subtitle)
