"""
Per-reconciliation support package workbook (.xlsx).

Cover sheet + Items sheet + (optional) Evidence sheet + (optional)
Notes sheet. Uses the same xlsx_builder named styles as Period Export
so the look is consistent across every download.
"""
from __future__ import annotations

import logging
import re
import uuid
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import desc, select
from sqlalchemy.ext.asyncio import AsyncSession

from models.reconciliation import (
    Reconciliation,
    ReconciliationItem,
    ReconNote,
    ReconTransaction,
)
from modules.exports.period_workbook import build_cover_sheet
from modules.exports.xlsx_builder import (
    add_sheet_title,
    fmt_date,
    freeze_header,
    register_styles,
    safe_dec,
    set_column_widths,
    write_row,
    write_table_header,
)

logger = logging.getLogger(__name__)


async def build_recon_workbook(
    *,
    db: AsyncSession,
    recon_id: uuid.UUID,
    company_name: str,
    generated_by_name: str,
) -> tuple[bytes, str]:
    """Build the per-recon support package + return (bytes, suggested_filename)."""
    recon = (await db.execute(
        select(Reconciliation).where(Reconciliation.id == recon_id)
    )).scalar_one_or_none()
    if recon is None:
        raise ValueError("Reconciliation not found")

    items = list((await db.execute(
        select(ReconciliationItem)
        .where(ReconciliationItem.reconciliation_id == recon_id)
        .order_by(desc(ReconciliationItem.subledger_balance))
    )).scalars().all())
    item_ids = [i.id for i in items]

    txns: list[ReconTransaction] = []
    if item_ids:
        txns = list((await db.execute(
            select(ReconTransaction)
            .where(ReconTransaction.reconciliation_item_id.in_(item_ids))
            .order_by(ReconTransaction.txn_date.desc().nullslast())
        )).scalars().all())

    notes = list((await db.execute(
        select(ReconNote)
        .where(ReconNote.reconciliation_id == recon_id)
        .order_by(ReconNote.created_at)
    )).scalars().all())

    name_by_item = {i.id: i.entity_name for i in items}

    wb = Workbook()
    register_styles(wb)
    if wb.active is not None:
        wb.remove(wb.active)

    build_cover_sheet(
        wb,
        company_name=company_name,
        package_title="Reconciliation Support Package",
        period_label=f"Period ending {fmt_date(recon.period_end)}",
        generated_by=generated_by_name,
        contents_text=(
            f"Reconciliation: {recon.name} · Type {recon.recon_type} · "
            f"GL ${float(recon.gl_total):,.2f} · Subledger "
            f"${float(recon.subledger_total):,.2f} · "
            f"Difference ${float(recon.difference):,.2f} · "
            f"Status {recon.status}"
        ),
        footer_text=(
            "Items are sorted by subledger balance descending. Evidence and Notes "
            "sheets are included only when present."
        ),
    )

    _build_items_sheet(wb, items)
    if txns:
        _build_evidence_sheet(wb, txns, name_by_item)
    if notes:
        _build_notes_sheet(wb, notes)
    if recon.ai_summary:
        _build_ai_summary_sheet(wb, recon.ai_summary)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe = re.sub(r"[^A-Za-z0-9 _-]+", "", recon.name or "").strip()
    safe = re.sub(r"\s+", "_", safe)[:50] or "reconciliation"
    fname = f"{safe}_{recon.period_end.isoformat()}_reconciliation.xlsx"

    return buf.read(), fname


def _build_items_sheet(wb: Workbook, items: list[ReconciliationItem]) -> None:
    ws = wb.create_sheet("Items")
    header_row = add_sheet_title(
        ws, "Reconciliation Items",
        subtitle="GL vs subledger by entity, with aging buckets",
    )
    headers = [
        "Entity", "GL Balance", "Subledger", "Difference",
        "Current", "1-30", "31-60", "61-90", "Over 90",
        "Risk", "Status", "AI Commentary",
    ]
    set_column_widths(ws, [28, 14, 14, 14, 12, 12, 12, 12, 12, 10, 12, 50])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    if not items:
        write_row(ws, header_row + 1, [
            ("No items on this reconciliation.", "nx_cell_muted"),
            *[("", "nx_cell_text") for _ in range(11)],
        ])
        return

    sums = {k: Decimal("0") for k in [
        "gl", "sl", "diff", "cur", "b1", "b2", "b3", "b4"
    ]}
    r = header_row + 1
    for it in items:
        gl  = safe_dec(it.gl_balance)
        sl  = safe_dec(it.subledger_balance)
        df  = safe_dec(it.difference)
        a0  = safe_dec(it.aging_current)
        a1  = safe_dec(it.aging_1_30)
        a2  = safe_dec(it.aging_31_60)
        a3  = safe_dec(it.aging_61_90)
        a4  = safe_dec(it.aging_over_90)
        sums["gl"]  += gl
        sums["sl"]  += sl
        sums["diff"]+= df
        sums["cur"] += a0
        sums["b1"]  += a1
        sums["b2"]  += a2
        sums["b3"]  += a3
        sums["b4"]  += a4

        write_row(ws, r, [
            (it.entity_name or "",                  "nx_cell_text"),
            (gl,                                    "nx_cell_money"),
            (sl,                                    "nx_cell_money"),
            (df,                                    "nx_cell_money"),
            (a0,                                    "nx_cell_money"),
            (a1,                                    "nx_cell_money"),
            (a2,                                    "nx_cell_money"),
            (a3,                                    "nx_cell_money"),
            (a4,                                    "nx_cell_money"),
            ((it.risk_level or "low").title(),      "nx_cell_text"),
            ((it.status or "pending").title(),      "nx_cell_text"),
            (it.ai_commentary or "",                "nx_cell_muted"),
        ])
        r += 1

    write_row(ws, r, [
        ("TOTAL", "nx_total_label"),
        (sums["gl"],  "nx_total_money"),
        (sums["sl"],  "nx_total_money"),
        (sums["diff"],"nx_total_money"),
        (sums["cur"], "nx_total_money"),
        (sums["b1"],  "nx_total_money"),
        (sums["b2"],  "nx_total_money"),
        (sums["b3"],  "nx_total_money"),
        (sums["b4"],  "nx_total_money"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
    ])


def _build_evidence_sheet(
    wb: Workbook,
    txns: list[ReconTransaction],
    name_by_item: dict[uuid.UUID, str],
) -> None:
    ws = wb.create_sheet("Evidence")
    header_row = add_sheet_title(
        ws, "Supporting Transactions",
        subtitle="Reconciling items pulled from QBO for evidence",
    )
    headers = ["Entity", "Category", "Type", "Number", "Date", "Amount", "Memo"]
    set_column_widths(ws, [28, 16, 14, 16, 12, 14, 50])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    sum_amt = Decimal("0")
    r = header_row + 1
    for t in txns:
        amt = safe_dec(t.amount)
        sum_amt += amt
        write_row(ws, r, [
            (name_by_item.get(t.reconciliation_item_id, ""), "nx_cell_text"),
            ((t.category or "").replace("_", " ").title(),   "nx_cell_text"),
            (t.txn_type or "",                               "nx_cell_text"),
            (t.txn_number or "",                             "nx_cell_text"),
            (t.txn_date if t.txn_date else "",               "nx_cell_date" if t.txn_date else "nx_cell_text"),
            (amt,                                            "nx_cell_money"),
            (t.memo or "",                                   "nx_cell_muted"),
        ])
        r += 1

    write_row(ws, r, [
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        ("", "nx_total_label"),
        ("TOTAL", "nx_total_label"),
        (sum_amt, "nx_total_money"),
        ("", "nx_total_label"),
    ])


def _build_notes_sheet(wb: Workbook, notes: list[ReconNote]) -> None:
    ws = wb.create_sheet("Notes")
    header_row = add_sheet_title(
        ws, "Reviewer Notes",
        subtitle="In chronological order",
    )
    headers = ["When (UTC)", "Note"]
    set_column_widths(ws, [22, 90])
    write_table_header(ws, header_row, headers)
    freeze_header(ws, header_row)

    r = header_row + 1
    for n in notes:
        write_row(ws, r, [
            (n.created_at.strftime("%Y-%m-%d %H:%M") if n.created_at else "", "nx_cell_text"),
            (n.body or "", "nx_cell_text"),
        ])
        r += 1


def _build_ai_summary_sheet(wb: Workbook, ai_summary: str) -> None:
    ws = wb.create_sheet("AI Summary")
    header_row = add_sheet_title(
        ws, "AI Summary",
        subtitle="Generated commentary on the reconciliation as a whole",
    )
    set_column_widths(ws, [110])
    ws.cell(row=header_row, column=1, value=ai_summary).style = "nx_cell_text"
    ws.row_dimensions[header_row].height = 240
