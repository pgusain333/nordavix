"""
Shared xlsx styling + helpers. Every sheet in every export uses the
same look — clean, no fills, no color. CPA-workpaper aesthetic:
black text on white with thin grey borders for the grid, thicker
bottom border on header rows, double top border for totals.

Used by:
  - modules/exports/period_workbook.py  (Period Export)
  - modules/exports/recon_workbook.py   (per-account / per-recon export)
  - modules/exports/flux_workbook.py    (variance analysis export)
  - modules/exports/schedules_workbook.py (per-schedule type export)
"""
from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    NamedStyle,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ── Neutral tokens (no color) ───────────────────────────────────────────
DARK  = "111827"   # body text — near-black
MUTED = "6B7280"   # captions / muted labels
GRID  = "D1D5DB"   # thin grid border
RULE  = "111827"   # thick rule (under headers, above totals)

# ── Border sides ────────────────────────────────────────────────────────
_THIN     = Side(style="thin",   color=GRID)
_MEDIUM   = Side(style="medium", color=RULE)
_DOUBLE   = Side(style="double", color=RULE)


def register_styles(wb: Workbook) -> None:
    """Register named styles on a fresh workbook so we can reference
    them by name from any sheet without redeclaring."""
    styles = [
        NamedStyle(
            name="nx_h1",  # cover-sheet H1
            font=Font(name="Calibri", size=24, bold=True, color=DARK),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        NamedStyle(
            name="nx_h2",  # sheet titles
            font=Font(name="Calibri", size=16, bold=True, color=DARK),
            alignment=Alignment(horizontal="left", vertical="center"),
        ),
        NamedStyle(
            name="nx_label",  # field labels on cover sheet
            font=Font(name="Calibri", size=10, bold=True, color=MUTED),
            alignment=Alignment(horizontal="left", vertical="top"),
        ),
        NamedStyle(
            name="nx_value",  # field values on cover sheet
            font=Font(name="Calibri", size=12, color=DARK),
            alignment=Alignment(horizontal="left", vertical="top"),
        ),
        NamedStyle(
            name="nx_table_head",  # column header row on table sheets
            font=Font(name="Calibri", size=10, bold=True, color=DARK),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=Border(top=_THIN, bottom=_MEDIUM, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_cell_text",
            font=Font(name="Calibri", size=10, color=DARK),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_cell_money",
            font=Font(name="Calibri", size=10, color=DARK),
            alignment=Alignment(horizontal="right", vertical="center"),
            number_format='_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            border=Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_cell_int",
            font=Font(name="Calibri", size=10, color=DARK),
            alignment=Alignment(horizontal="right", vertical="center"),
            number_format='#,##0',
            border=Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_cell_pct",
            font=Font(name="Calibri", size=10, color=DARK),
            alignment=Alignment(horizontal="right", vertical="center"),
            number_format='0.0%',
            border=Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_cell_date",
            font=Font(name="Calibri", size=10, color=DARK),
            alignment=Alignment(horizontal="left", vertical="center"),
            number_format='mm-dd-yyyy',
            border=Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_cell_muted",
            font=Font(name="Calibri", size=10, italic=True, color=MUTED),
            alignment=Alignment(horizontal="left", vertical="center", wrap_text=True),
            border=Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_total_label",
            font=Font(name="Calibri", size=11, bold=True, color=DARK),
            alignment=Alignment(horizontal="right", vertical="center"),
            border=Border(top=_DOUBLE, bottom=_THIN, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_total_money",
            font=Font(name="Calibri", size=11, bold=True, color=DARK),
            alignment=Alignment(horizontal="right", vertical="center"),
            number_format='_-$* #,##0.00_-;-$* #,##0.00_-;_-$* "-"??_-;_-@_-',
            border=Border(top=_DOUBLE, bottom=_THIN, left=_THIN, right=_THIN),
        ),
        NamedStyle(
            name="nx_total_int",
            font=Font(name="Calibri", size=11, bold=True, color=DARK),
            alignment=Alignment(horizontal="right", vertical="center"),
            number_format='#,##0',
            border=Border(top=_DOUBLE, bottom=_THIN, left=_THIN, right=_THIN),
        ),
    ]
    for s in styles:
        if s.name not in wb.named_styles:
            wb.add_named_style(s)


# ── Sheet helpers ───────────────────────────────────────────────────────

def add_sheet_title(ws: Worksheet, title: str, *, subtitle: str | None = None) -> int:
    """Write the title at A1; subtitle at A2 (italic muted). Returns the
    row number where the header row should start (skipping a blank row)."""
    ws["A1"] = title
    ws["A1"].style = "nx_h2"
    ws.row_dimensions[1].height = 28
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(name="Calibri", size=10, italic=True, color=MUTED)
        return 4  # leave row 3 blank
    return 3


def write_table_header(ws: Worksheet, row: int, headers: list[str]) -> None:
    """Write a single styled header row. Caller supplies the column
    widths separately via set_column_widths."""
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.style = "nx_table_head"
    ws.row_dimensions[row].height = 26


def set_column_widths(ws: Worksheet, widths: list[float]) -> None:
    """Set explicit widths (in Excel character units). Most sheets need
    this — openpyxl's auto-fit isn't great."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_row(
    ws: Worksheet, row: int, cells: list[tuple[Any, str]],
) -> None:
    """Write one data row. Each cell is (value, style_name).
    Pass empty string + 'nx_cell_text' for blank cells to keep borders."""
    for col, (value, style_name) in enumerate(cells, start=1):
        # Convert Decimal to float for openpyxl (it can serialize Decimal
        # but the number_format applies more reliably on float).
        if isinstance(value, Decimal):
            value = float(value)
        cell = ws.cell(row=row, column=col, value=value)
        cell.style = style_name


def freeze_header(ws: Worksheet, header_row: int) -> None:
    """Freeze rows above (and including) the header so the table
    header stays visible as the user scrolls."""
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


# ── Formatting helpers callable from anywhere ──────────────────────────

def fmt_date(d: date | datetime | str | None) -> str:
    if d is None:
        return ""
    if isinstance(d, str):
        return d
    if isinstance(d, datetime):
        return d.strftime("%m-%d-%Y")
    return d.strftime("%m-%d-%Y")


def safe_dec(value: Any) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        return Decimal(str(value))
    except Exception:
        return Decimal("0")
