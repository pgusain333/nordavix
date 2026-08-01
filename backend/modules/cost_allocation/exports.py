"""Nordavix Allocate — the export engine.

Two deliverables, one dataset:

  • an Excel workbook — the working file. Every sheet a reviewer or an examiner
    would ask for, cross-footed, with the basis behind each number rather than
    only the number.
  • a PDF report — the client deliverable. What was done, on what authority,
    what it concluded, and what reaches the return.

Both are assembled from `assemble()`, so the two can't tell different stories
about the same year. That matters more here than it looks: the workbook is what
gets defended and the PDF is what gets circulated, and a firm that hands over a
report which doesn't tie to its own workpaper has a problem no formatting fixes.

Every figure crosses in as a Decimal string and is written as a NUMBER in Excel
(so the client can foot it) but never re-derived — the arithmetic already
happened in the engine, and doing it twice is how two answers appear.
"""
from __future__ import annotations

import io
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from models.cost_allocation import (
    AllocEligibility,
    AllocEmployee,
    AllocPool,
    AllocRun,
    AllocSettings,
    AllocSpace,
)
from modules.cost_allocation.annual import build_annual
from modules.cost_allocation.engine import is_effective, normalize_frequency

# ── Assembly ──────────────────────────────────────────────────────────────────


async def assemble(
    db: AsyncSession, *, tenant_id: uuid.UUID, tax_year: int, client_name: str,
) -> dict[str, Any]:
    """Everything both exports need, read once.

    The registries are read AS OF the year end, which is the state the year was
    allocated on. Reading them as of today would produce a workpaper describing
    a configuration the numbers were never computed from.
    """
    annual = await build_annual(db, tenant_id=tenant_id, tax_year=tax_year)
    as_of = date.fromisoformat(annual["year_end"])

    cfg = (await db.execute(select(AllocSettings))).scalars().first()
    pools = list((await db.execute(
        select(AllocPool).order_by(AllocPool.sort_order, AllocPool.name)
    )).scalars().all())
    spaces = [
        s for s in (await db.execute(select(AllocSpace).order_by(AllocSpace.name))).scalars().all()
        if is_effective(s.effective_from, s.effective_to, as_of)
    ]
    employees = [
        e for e in (await db.execute(
            select(AllocEmployee).order_by(AllocEmployee.name)
        )).scalars().all()
        if e.active and is_effective(e.effective_from, e.effective_to, as_of)
    ]
    elig = (await db.execute(
        select(AllocEligibility).where(AllocEligibility.tax_year == tax_year)
    )).scalars().first()
    year_start = date.fromisoformat(annual["year_start"])
    runs = list((await db.execute(
        select(AllocRun)
        .where(AllocRun.period_end >= year_start, AllocRun.period_end <= as_of)
        .order_by(AllocRun.period_end)
    )).scalars().all())

    return {
        "client_name": client_name,
        "tax_year": tax_year,
        "prepared_at": datetime.now(UTC),
        "annual": annual,
        "settings": {
            "method": (cfg.method if cfg else "books_records"),
            "has_afs": bool(cfg.has_afs) if cfg else False,
            "inventory_method": (cfg.inventory_method if cfg else "rollforward"),
            "fiscal_year_end": (cfg.fiscal_year_end if cfg else None),
            "frequency": normalize_frequency(cfg.allocation_frequency if cfg else None),
            "election_attested_at": (
                cfg.election_attested_at.isoformat()
                if cfg and cfg.election_attested_at else None
            ),
        },
        "pools": [{
            "name": p.name, "treatment": p.treatment, "driver": p.driver,
            "blend_payroll_wt": _s(p.blend_payroll_wt),
            "blend_occupancy_wt": _s(p.blend_occupancy_wt),
            "fixed_pct": _s(p.fixed_pct),
            "form_1125a_line": p.form_1125a_line or "other",
            "active": p.active, "notes": p.notes,
        } for p in pools],
        "spaces": [{
            "name": s.name, "function": s.function,
            "square_feet": _s(s.square_feet),
            "production_pct": _s(s.production_pct),
            "effective_from": s.effective_from.isoformat(),
        } for s in spaces],
        "employees": [{
            "name": e.name, "department": e.department, "job_title": e.job_title,
            "function": e.function, "production_pct": _s(e.production_pct),
            "effective_from": e.effective_from.isoformat(),
        } for e in employees],
        "eligibility": None if elig is None else {
            "tax_year": elig.tax_year,
            "threshold": _s(elig.threshold),
            "three_year_avg": _s(elig.three_year_avg),
            "eligible": elig.eligible,
            "has_afs": elig.has_afs,
            "method_available": elig.method_available,
            "reason": elig.reason,
            "aggregation_note": elig.aggregation_note,
            "entities": elig.entities or [],
            "tested_at": elig.tested_at.isoformat() if elig.tested_at else None,
        },
        "runs": [{
            "period_start": r.period_start.isoformat(),
            "period_end": r.period_end.isoformat(),
            "status": r.status,
            "payroll_factor": _s(r.payroll_factor),
            "occupancy_factor": _s(r.occupancy_factor),
            "driver_basis": r.driver_basis,
            "posted_doc_number": r.posted_doc_number,
            "posted_at": r.posted_at.isoformat() if r.posted_at else None,
            "approved_at": r.approved_at.isoformat() if r.approved_at else None,
        } for r in runs if _in_year(r.period_end, annual)],
    }


def _s(v: Any) -> str | None:
    return str(v) if v is not None else None


def _in_year(period_end: date, annual: dict) -> bool:
    return annual["year_start"] <= period_end.isoformat() <= annual["year_end"]


def _num(v: Any) -> float | str | None:
    """Decimal string → float for Excel, so the client can foot the column.

    Display formatting is a cell format, never a pre-formatted string: a number
    stored as text is a number nobody can add up.
    """
    if v is None or v == "":
        return None
    try:
        return float(Decimal(str(v)))
    except (InvalidOperation, ValueError):
        return str(v)


# ── Excel workbook ────────────────────────────────────────────────────────────

_MONEY = "#,##0.00;(#,##0.00)"
_PCT = "0.0000%"


def build_workbook(data: dict[str, Any]) -> bytes:
    """The working file — one sheet per question a reviewer will ask."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    INK = "FF14181A"
    GREEN = "FF3E8F66"
    MUTED = "FF8A8F98"
    HEAD_BG = "FFEAF4EE"
    RULE = Side(style="thin", color="FFE6E4DF")

    wb = Workbook()
    annual = data["annual"]
    check = annual["checklist"]

    def sheet(title: str):
        ws = wb.create_sheet(title[:31])
        ws.sheet_view.showGridLines = False
        return ws

    def title_row(ws, text: str, sub: str = "") -> int:
        ws["A1"] = text
        ws["A1"].font = Font(name="Calibri", size=14, bold=True, color=INK)
        if sub:
            ws["A2"] = sub
            ws["A2"].font = Font(name="Calibri", size=9, color=MUTED)
        return 4

    def header(ws, row: int, labels: list[str], widths: list[int]) -> int:
        for i, (label, w) in enumerate(zip(labels, widths, strict=True), start=1):
            c = ws.cell(row=row, column=i, value=label)
            c.font = Font(name="Calibri", size=9, bold=True, color=INK)
            c.fill = PatternFill("solid", fgColor=HEAD_BG)
            c.border = Border(bottom=RULE)
            c.alignment = Alignment(vertical="center")
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = ws.cell(row=row + 1, column=1)
        return row + 1

    def put(ws, row: int, values: list[Any], *, fmts: dict[int, str] | None = None,
            bold: bool = False) -> int:
        for i, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=i, value=v)
            c.font = Font(name="Calibri", size=10, bold=bold, color=INK)
            if fmts and i in fmts:
                c.number_format = fmts[i]
                c.alignment = Alignment(horizontal="right")
        return row + 1

    # ── Cover ────────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Section 471(c) cost allocation"
    ws["A1"].font = Font(name="Calibri", size=18, bold=True, color=INK)
    ws["A2"] = data["client_name"]
    ws["A2"].font = Font(name="Calibri", size=12, color=GREEN, bold=True)
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 52

    s = data["settings"]
    facts = [
        ("Tax year", str(data["tax_year"])),
        ("Period", f"{annual['year_start']} to {annual['year_end']}"),
        ("Allocation performed", "Annually" if s["frequency"] == "annual" else "Monthly"),
        ("Method", "Conform to AFS — §471(c)(1)(B)(i)" if s["method"] == "afs"
                   else "Books and records — §471(c)(1)(B)(ii)"),
        ("Applicable financial statement", "Yes" if s["has_afs"] else "No"),
        ("Inventory method", "Period roll-forward"),
        ("Fiscal year end", s["fiscal_year_end"] or "12-31 (calendar)"),
        ("Prepared", data["prepared_at"].date().isoformat()),
        ("", ""),
        ("COMPLETENESS", ""),
        ("Complete", "Yes" if annual["complete"] else "No"),
        ("Periods expected", check["months_expected"]),
        ("Periods present", check["months_present"]),
        ("Missing", ", ".join(check["missing_periods"]) or "None"),
        ("Not approved", ", ".join(check["unapproved_periods"]) or "None"),
        ("Not confirmed posted", ", ".join(check["unposted_periods"]) or "None"),
        ("Inventory chain breaks", str(len(check["inventory_breaks"])) if check["inventory_breaks"] else "None"),
        ("§448(c) test concluded", "Yes" if check["eligibility_concluded"] else "No"),
    ]
    r = 4
    for label, value in facts:
        ws.cell(row=r, column=1, value=label).font = Font(
            name="Calibri", size=10, bold=label.isupper(), color=MUTED if not label.isupper() else INK,
        )
        ws.cell(row=r, column=2, value=value).font = Font(name="Calibri", size=10, color=INK)
        r += 1

    r += 1
    ws.cell(row=r, column=1, value=(
        "Line 4 of Form 1125-A (additional §263A costs) is not presented: §280E "
        "denies §263A to this taxpayer, which is the reason §471(c) is used."
    )).font = Font(name="Calibri", size=9, italic=True, color=MUTED)

    # ── Summary ──────────────────────────────────────────────────────────────
    ws = sheet("Summary")
    r = title_row(ws, "Annual summary", f"{data['client_name']} · tax year {data['tax_year']}")
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18

    t = annual["totals"]
    for label, value in (
        ("Total expenses in scope", t["total_expenses"]),
        ("Capitalized into inventory", t["capitalized"]),
        ("Disallowed under §280E", t["disallowed"]),
    ):
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, color=INK)
        c = ws.cell(row=r, column=2, value=_num(value))
        c.number_format = _MONEY
        c.font = Font(name="Calibri", size=10, bold=True, color=INK)
        r += 1

    rf = annual["roll_forward"]
    r += 1
    ws.cell(row=r, column=1, value="INVENTORY ROLL-FORWARD").font = Font(
        name="Calibri", size=10, bold=True, color=INK)
    r += 1
    if rf:
        for label, value, bold in (
            ("Beginning inventory", rf["beginning_inventory"], False),
            ("Capitalized cost", rf["capitalized"], False),
            ("Purchases", rf["purchases"], False),
            ("Ending inventory", rf["ending_inventory"], False),
            ("Cost of goods sold", rf["cogs"], True),
        ):
            ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, color=INK)
            c = ws.cell(row=r, column=2, value=_num(value))
            c.number_format = _MONEY
            c.font = Font(name="Calibri", size=10, bold=bold, color=INK)
            r += 1
    else:
        ws.cell(row=r, column=1, value="Not available — inventory was not captured").font = Font(
            name="Calibri", size=10, italic=True, color=MUTED)
        r += 1

    f = annual["form_1125a"]
    r += 1
    ws.cell(row=r, column=1, value="FORM 1125-A").font = Font(
        name="Calibri", size=10, bold=True, color=INK)
    r += 1
    for label, value, bold in (
        ("Line 1  Inventory at beginning of year", f["line_1_beginning_inventory"], False),
        ("Line 2  Purchases", f["line_2_purchases"], False),
        ("Line 3  Cost of labor", f["line_3_cost_of_labor"], False),
        ("Line 5  Other costs", f["line_5_other_costs"], False),
        ("Line 6  Total", f["line_6_total"], True),
        ("Line 7  Inventory at end of year", f["line_7_ending_inventory"], False),
        ("Line 8  Cost of goods sold", f["line_8_cogs"], True),
    ):
        ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, color=INK)
        c = ws.cell(row=r, column=2, value=_num(value))
        c.number_format = _MONEY
        c.font = Font(name="Calibri", size=10, bold=bold, color=INK)
        r += 1

    # ── Periods ──────────────────────────────────────────────────────────────
    ws = sheet("Periods")
    r = title_row(ws, "Period roll", "Every period in the year, as concluded")
    r = header(ws, r, [
        "Period end", "Status", "Posted", "Doc no.", "Payroll factor", "Occupancy factor",
        "Total expenses", "Capitalized", "Disallowed", "Beginning inv.", "Purchases", "Ending inv.",
    ], [13, 12, 9, 12, 15, 16, 15, 15, 15, 15, 13, 15])

    runs_by_pe = {r_["period_end"]: r_ for r_ in data["runs"]}
    money_cols = {7: _MONEY, 8: _MONEY, 9: _MONEY, 10: _MONEY, 11: _MONEY, 12: _MONEY}
    pct_cols = {5: _PCT, 6: _PCT}
    for m in annual["months"]:
        run = runs_by_pe.get(m["period_end"], {})
        r = put(ws, r, [
            m["period_end"], m["status"], "Yes" if m["posted"] else "No",
            run.get("posted_doc_number") or "",
            _num(run.get("payroll_factor")), _num(run.get("occupancy_factor")),
            _num(m["total_expenses"]), _num(m["capitalized"]), _num(m["disallowed"]),
            _num(m["beginning_inventory"]), _num(m["purchases"]), _num(m["ending_inventory"]),
        ], fmts={**money_cols, **pct_cols})
    r = put(ws, r, [
        "TOTAL", "", "", "", None, None,
        _num(t["total_expenses"]), _num(t["capitalized"]), _num(t["disallowed"]),
    ], fmts=money_cols, bold=True)

    # ── Accounts ─────────────────────────────────────────────────────────────
    ws = sheet("Accounts")
    r = title_row(ws, "Account detail", "Every expense account, for the year")
    r = header(ws, r, [
        "Account no.", "Account", "Pool", "Treatment", "Gross", "Capitalized", "Disallowed",
    ], [13, 42, 24, 12, 15, 15, 15])
    acc_fmts = {5: _MONEY, 6: _MONEY, 7: _MONEY}
    for a in annual["by_account"]:
        r = put(ws, r, [
            a["account_number"] or "", a["account_name"] or a["qbo_account_id"],
            a["pool_name"], a["treatment"],
            _num(a["gross"]), _num(a["capitalized"]), _num(a["disallowed"]),
        ], fmts=acc_fmts)
    if annual["by_account"]:
        r = put(ws, r, [
            "", "TOTAL", "", "",
            _num(sum(Decimal(a["gross"]) for a in annual["by_account"])),
            _num(t["capitalized"]), _num(t["disallowed"]),
        ], fmts=acc_fmts, bold=True)

    # ── Pools ────────────────────────────────────────────────────────────────
    ws = sheet("Pools")
    r = title_row(ws, "Cost pools", "How each group of cost is treated, and on what driver")
    r = header(ws, r, [
        "Pool", "Treatment", "Driver", "Payroll wt %", "Occupancy wt %", "Fixed %",
        "Form 1125-A", "Capitalized", "Active",
    ], [26, 12, 12, 13, 15, 10, 20, 15, 8])
    cap_by_pool = {p["pool_name"]: p["capitalized"] for p in annual["by_pool"]}
    for p in data["pools"]:
        r = put(ws, r, [
            p["name"], p["treatment"], p["driver"] or "",
            _num(p["blend_payroll_wt"]), _num(p["blend_occupancy_wt"]), _num(p["fixed_pct"]),
            "Line 3 — cost of labor" if p["form_1125a_line"] == "labor" else "Line 5 — other costs",
            _num(cap_by_pool.get(p["name"])),
            "Yes" if p["active"] else "No",
        ], fmts={8: _MONEY})

    # ── Drivers ──────────────────────────────────────────────────────────────
    ws = sheet("Drivers")
    r = title_row(ws, "Allocation drivers",
                  f"Square footage and payroll classifications as at {annual['year_end']}")
    ws.cell(row=r, column=1, value="SPACES").font = Font(name="Calibri", size=10, bold=True, color=INK)
    r += 1
    r = header(ws, r, ["Space", "Function", "Square feet", "Production %", "Effective from"],
               [30, 16, 13, 14, 14])
    for sp in data["spaces"]:
        r = put(ws, r, [
            sp["name"], sp["function"], _num(sp["square_feet"]),
            _num(sp["production_pct"]), sp["effective_from"],
        ], fmts={3: "#,##0.00", 4: "0.00"})

    r += 2
    ws.cell(row=r, column=1, value="EMPLOYEES").font = Font(name="Calibri", size=10, bold=True, color=INK)
    r += 1
    r = header(ws, r, ["Employee", "Department", "Job title", "Function", "Production %", "Effective from"],
               [26, 20, 22, 16, 14, 14])
    for e in data["employees"]:
        r = put(ws, r, [
            e["name"], e["department"] or "", e["job_title"] or "", e["function"],
            _num(e["production_pct"]), e["effective_from"],
        ], fmts={5: "0.00"})

    # ── Eligibility ──────────────────────────────────────────────────────────
    ws = sheet("Eligibility")
    r = title_row(ws, "§448(c) small business taxpayer test",
                  "Gross receipts aggregate across commonly controlled entities — §52(a)/(b), §414(m)/(o)")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 22
    e = data["eligibility"]
    if e is None:
        ws.cell(row=r, column=1, value="No conclusion on file for this tax year.").font = Font(
            name="Calibri", size=10, italic=True, color=MUTED)
    else:
        for label, value, fmt in (
            ("Three-year average gross receipts", _num(e["three_year_avg"]), _MONEY),
            ("Threshold applied", _num(e["threshold"]), _MONEY),
            ("Conclusion", "Eligible" if e["eligible"] else "NOT eligible", None),
            ("Applicable financial statement", "Yes" if e["has_afs"] else "No", None),
            ("Prong available", e["method_available"] or "—", None),
            ("Tested", (e["tested_at"] or "")[:10], None),
        ):
            ws.cell(row=r, column=1, value=label).font = Font(name="Calibri", size=10, color=MUTED)
            c = ws.cell(row=r, column=2, value=value)
            c.font = Font(name="Calibri", size=10, bold=True, color=INK)
            if fmt:
                c.number_format = fmt
            r += 1
        if e["aggregation_note"]:
            r += 1
            ws.cell(row=r, column=1, value="Aggregation basis").font = Font(
                name="Calibri", size=10, color=MUTED)
            ws.cell(row=r, column=2, value=e["aggregation_note"]).font = Font(
                name="Calibri", size=10, color=INK)
            r += 1
        r += 1
        r = header(ws, r, ["Entity", "Year", "Gross receipts", "Source"], [34, 10, 18, 14])
        for row in e["entities"]:
            r = put(ws, r, [
                row.get("entity", ""), row.get("year"), _num(row.get("amount")),
                row.get("source", "manual"),
            ], fmts={3: _MONEY})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
