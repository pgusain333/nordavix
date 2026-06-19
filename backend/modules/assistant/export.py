"""Export a single NDVX Copilot answer to a branded PDF or Excel file.

Per-answer export: the answer's text (light Markdown) plus any charts the copilot
rendered. PDF uses ReportLab (monochrome + green accent, matching the app's other
exports); Excel uses openpyxl with a native chart per chart-spec. Pure formatting
— no DB, no AI — so it's cheap and safe to call from the export endpoint.
"""
from __future__ import annotations

import io
import re
from datetime import date

from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

_GREEN = colors.HexColor("#0F6E56")
_INK = colors.HexColor("#1A1A1A")
_MUTED = colors.HexColor("#6B6B6B")
_RULE = colors.HexColor("#D9D9D9")
_PIE_COLORS = [
    colors.HexColor(c)
    for c in ("#1D9E75", "#5DCAA5", "#888780", "#B4B2A9", "#0F6E56", "#D3D1C7")
]

_CELL = ParagraphStyle("cell", fontName="Helvetica", fontSize=9, leading=12, textColor=_INK)


def _esc(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _inline_md(s: str) -> str:
    s = _esc(s)
    s = re.sub(r"\*\*(.+?)\*\*", r"<b>\1</b>", s)
    s = re.sub(r"`(.+?)`", r'<font face="Courier">\1</font>', s)
    return s


def _num(v, unit: str = "") -> str:
    try:
        n = float(v)
    except (TypeError, ValueError):
        return str(v)
    if unit == "%":
        return f"{round(n, 1)}%"
    return f"{unit}{n:,.2f}" if unit else f"{n:,.2f}"


# ── PDF ──────────────────────────────────────────────────────────────────────

def _pipe_table(lines: list[str]) -> Table | None:
    rows: list[list[str]] = []
    for ln in lines:
        if set(ln) <= set(" |:-"):  # separator row (---|---)
            continue
        rows.append([c.strip() for c in ln.strip().strip("|").split("|")])
    if not rows:
        return None
    ncol = max(len(r) for r in rows)
    body = [[Paragraph(_inline_md(c), _CELL) for c in (r + [""] * (ncol - len(r)))] for r in rows]
    t = Table(body, hAlign="LEFT")
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("LINEBELOW", (0, 0), (-1, 0), 1, _GREEN),
        ("LINEBELOW", (0, 1), (-1, -1), 0.5, _RULE),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    return t


def _answer_flowables(answer: str, body, bullet, h2) -> list:
    out: list = []
    for block in re.split(r"\n\s*\n", (answer or "").strip()):
        lines = [ln.rstrip() for ln in block.split("\n") if ln.strip()]
        if not lines:
            continue
        if len(lines) >= 2 and all("|" in ln for ln in lines):
            tbl = _pipe_table(lines)
            if tbl is not None:
                out.append(tbl)
                out.append(Spacer(1, 6))
                continue
        if lines[0].lstrip().startswith("#"):
            out.append(Paragraph(_inline_md(lines[0].lstrip("# ").strip()), h2))
            lines = lines[1:]
            if not lines:
                continue
        if all(re.match(r"^\s*([-*]|\d+\.)\s+", ln) for ln in lines):
            for ln in lines:
                out.append(Paragraph(_inline_md(re.sub(r"^\s*([-*]|\d+\.)\s+", "", ln)), bullet, bulletText="•"))
            continue
        out.append(Paragraph(_inline_md(" ".join(lines)), body))
    return out


def _chart_drawing(chart: dict):
    from reportlab.graphics.shapes import Drawing
    data = chart.get("data") or []
    if not data:
        return None
    try:
        vals = [float(p.get("value")) for p in data]
        labels = [str(p.get("label"))[:12] for p in data]
        ctype = chart.get("type")
        if ctype == "pie":
            from reportlab.graphics.charts.piecharts import Pie
            d = Drawing(440, 190)
            pie = Pie()
            pie.x, pie.y, pie.width, pie.height = 30, 15, 160, 160
            pie.data = [max(0.0, v) for v in vals] or [1]
            pie.labels = labels
            pie.sideLabels = True
            for i in range(len(pie.data)):
                pie.slices[i].fillColor = _PIE_COLORS[i % len(_PIE_COLORS)]
            d.add(pie)
            return d
        if ctype == "line":
            from reportlab.graphics.charts.linecharts import HorizontalLineChart
            d = Drawing(440, 190)
            lc = HorizontalLineChart()
            lc.x, lc.y, lc.width, lc.height = 40, 30, 380, 140
            lc.data = [vals]
            lc.categoryAxis.categoryNames = labels
            lc.lines[0].strokeColor = _GREEN
            lc.lines[0].strokeWidth = 2
            d.add(lc)
            return d
        from reportlab.graphics.charts.barcharts import VerticalBarChart
        d = Drawing(440, 190)
        bc = VerticalBarChart()
        bc.x, bc.y, bc.width, bc.height = 40, 35, 380, 140
        bc.data = [vals]
        bc.categoryAxis.categoryNames = labels
        bc.bars[0].fillColor = _GREEN
        bc.valueAxis.valueMin = min(0.0, *vals)
        d.add(bc)
        return d
    except Exception:
        return None


def _chart_table(chart: dict) -> Table:
    unit = chart.get("unit") or ""
    rows = [
        [Paragraph(_esc(str(p.get("label"))), _CELL), Paragraph(_num(p.get("value"), unit), _CELL)]
        for p in (chart.get("data") or [])
    ]
    t = Table(rows or [[Paragraph("—", _CELL)]], colWidths=[3.4 * inch, 1.6 * inch], hAlign="LEFT")
    t.setStyle(TableStyle([
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, _RULE),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TEXTCOLOR", (1, 0), (1, -1), _GREEN),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    return t


def build_answer_pdf(*, question: str, answer: str, charts: list[dict], company: str | None = None) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=0.8 * inch, bottomMargin=0.7 * inch,
        leftMargin=0.8 * inch, rightMargin=0.8 * inch,
        title="NDVX Copilot answer",
    )
    ss = getSampleStyleSheet()
    brand = ParagraphStyle("brand", parent=ss["Title"], fontName="Helvetica-Bold", fontSize=18, textColor=_INK, spaceAfter=2, alignment=0)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontName="Helvetica", fontSize=9, textColor=_MUTED, spaceAfter=10)
    qst = ParagraphStyle("q", parent=ss["Normal"], fontName="Helvetica-Bold", fontSize=12, textColor=_GREEN, spaceBefore=4, spaceAfter=8)
    body = ParagraphStyle("body", parent=ss["Normal"], fontName="Helvetica", fontSize=10, leading=15, textColor=_INK, spaceAfter=8)
    bullet = ParagraphStyle("bullet", parent=body, leftIndent=14, bulletIndent=2, spaceAfter=3)
    h2 = ParagraphStyle("h2", parent=ss["Normal"], fontName="Helvetica-Bold", fontSize=11, textColor=_INK, spaceBefore=10, spaceAfter=6)

    story: list = [
        Paragraph(_esc(company or "Nordavix"), brand),
        Paragraph(f"NDVX Copilot · generated {date.today().isoformat()}", sub),
        HRFlowable(width="100%", thickness=1, color=_RULE, spaceAfter=10),
    ]
    if question:
        story.append(Paragraph(_esc(question), qst))
    story.extend(_answer_flowables(answer, body, bullet, h2))
    for ch in (charts or []):
        story.append(Spacer(1, 8))
        if ch.get("title"):
            story.append(Paragraph(_esc(ch["title"]), h2))
        drawing = _chart_drawing(ch)
        if drawing is not None:
            story.append(drawing)
        story.append(_chart_table(ch))
    doc.build(story)
    return buf.getvalue()


# ── Excel ────────────────────────────────────────────────────────────────────

def _safe_sheet_title(title: str, fallback: str) -> str:
    t = re.sub(r"[\[\]:*?/\\]", " ", (title or "").strip())[:28]
    return t or fallback


def build_answer_xlsx(*, question: str, answer: str, charts: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Alignment, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "Answer"
    ws["A1"] = "NDVX Copilot"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Generated {date.today().isoformat()}"
    ws["A2"].font = Font(size=9, color="808080")
    row = 4
    if question:
        ws.cell(row, 1, "Question").font = Font(bold=True)
        ws.cell(row + 1, 1, question).alignment = Alignment(wrap_text=True, vertical="top")
        row += 3
    ws.cell(row, 1, "Answer").font = Font(bold=True)
    plain = re.sub(r"[*`#]", "", answer or "").strip()
    ws.cell(row + 1, 1, plain).alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 95

    for i, ch in enumerate(charts or [], start=1):
        cs = wb.create_sheet(title=_safe_sheet_title(ch.get("title"), f"Chart {i}"))
        cs["A1"] = "Label"
        cs["B1"] = "Value"
        cs["A1"].font = cs["B1"].font = Font(bold=True)
        data = ch.get("data") or []
        for j, p in enumerate(data, start=2):
            cs.cell(j, 1, str(p.get("label")))
            try:
                cs.cell(j, 2, float(p.get("value")))
            except (TypeError, ValueError):
                cs.cell(j, 2, 0)
        cs.column_dimensions["A"].width = 34
        n = len(data)
        if n:
            ctype = ch.get("type")
            chart = PieChart() if ctype == "pie" else LineChart() if ctype == "line" else BarChart()
            chart.title = ch.get("title") or "Chart"
            chart.height = 8
            chart.width = 15
            vals = Reference(cs, min_col=2, min_row=1, max_row=1 + n)
            cats = Reference(cs, min_col=1, min_row=2, max_row=1 + n)
            chart.add_data(vals, titles_from_data=True)
            chart.set_categories(cats)
            cs.add_chart(chart, "D2")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
